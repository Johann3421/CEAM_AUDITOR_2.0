"""Fichas Producto REST endpoints."""
import statistics
from collections import defaultdict
from datetime import datetime, timezone
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.db.database import get_db

router = APIRouter(prefix="/fichas", tags=["Fichas Producto"])

_TABLE = "fichas_producto"


def _safe_col(db: Session) -> list[str]:
    """Return actual column names from the DB table (may not exist yet)."""
    try:
        rows = db.execute(
            text(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_name = :t ORDER BY ordinal_position"
            ),
            {"t": _TABLE},
        ).fetchall()
        return [r[0] for r in rows]
    except Exception:
        return []


def _build_fichas_where(
    col_set: set,
    acuerdo_marco=None,
    catalogo=None,
    categoria=None,
    marca=None,
    estado=None,
    search=None,
    con_precio=None,
) -> tuple:
    """Build WHERE clause and params dict for fichas_producto queries."""
    filters: list[str] = []
    params: dict = {}

    def _f(col: str, val, param: str):
        if col in col_set and val:
            filters.append(f'"{col}" ILIKE :{param}')
            params[param] = f"%{val}%"

    _f("acuerdo_marco", acuerdo_marco, "acuerdo")
    _f("catálogo", catalogo, "catalogo")

    # categoria column may be stored as "categoría" (with accent) or "categora" (scraped variant)
    if categoria:
        cat_col = next((c for c in ("categoría", "categora") if c in col_set), None)
        if cat_col:
            filters.append(f'"{cat_col}" ILIKE :categoria')
            params["categoria"] = f"%{categoria}%"

    _f("marca", marca, "marca")
    _f("estado_ficha_producto", estado, "estado")

    if con_precio and "precio_referencia" in col_set:
        filters.append('"precio_referencia" IS NOT NULL')

    if search:
        search_cols = []
        for c in ("descripción_fichaproducto", "nro_parte_o_código_único_de_identificación",
                  "descripcin_fichaproducto", "nro_parte_o_cdigo_nico_de_identificacin"):
            if c in col_set:
                search_cols.append(f'"{c}" ILIKE :search')
        if search_cols:
            filters.append("(" + " OR ".join(search_cols) + ")")
            params["search"] = f"%{search}%"

    where_clause = ("WHERE " + " AND ".join(filters)) if filters else ""
    return filters, params, where_clause


@router.get("/")
def list_fichas(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    acuerdo_marco: Optional[str] = Query(None),
    catalogo: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    con_precio: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
):
    """List fichas-producto with optional filters and pagination."""
    cols = _safe_col(db)
    if not cols:
        return []

    col_set = set(cols)
    _, params, where_clause = _build_fichas_where(
        col_set, acuerdo_marco, catalogo, categoria, marca, estado, search, con_precio
    )
    params["skip"] = skip
    params["limit"] = limit
    quoted_cols = ", ".join(f'"{c}"' for c in cols)

    # Total count with the same filters (without limit/skip)
    count_params = {k: v for k, v in params.items() if k not in ("skip", "limit")}
    total_count = 0
    try:
        total_count = db.execute(
            text(f'SELECT COUNT(*) FROM {_TABLE} {where_clause}'),
            count_params,
        ).scalar() or 0
    except Exception:
        pass

    sql = text(
        f'SELECT {quoted_cols} FROM {_TABLE} {where_clause}'
        f' ORDER BY fecha_extraccion DESC NULLS LAST'
        f' LIMIT :limit OFFSET :skip'
    )
    rows = db.execute(sql, params).fetchall()
    return {
        "total": total_count,
        "items": [dict(zip(cols, row)) for row in rows],
    }


@router.get("/export")
def export_fichas_excel(
    acuerdo_marco: Optional[str] = Query(None),
    catalogo: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    con_precio: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
):
    """Export fichas producto as a styled Excel file respecting active filters."""
    import io
    from datetime import datetime
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    cols = _safe_col(db)
    if not cols:
        raise HTTPException(status_code=404, detail="No hay datos de fichas")

    col_set = set(cols)
    _, params, where_clause = _build_fichas_where(
        col_set, acuerdo_marco, catalogo, categoria, marca, estado, search, con_precio
    )
    params["limit"] = 100_000
    params["skip"] = 0

    quoted_cols = ", ".join(f'"{c}"' for c in cols)
    sql = text(
        f"SELECT {quoted_cols} FROM {_TABLE} {where_clause}"
        f" ORDER BY fecha_extraccion DESC NULLS LAST"
        f" LIMIT :limit OFFSET :skip"
    )
    try:
        rows = db.execute(sql, params).fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al consultar fichas: {e}")

    # ── Friendly column labels ─────────────────────────────────────────────
    COL_LABELS = {
        "nro_parte_o_código_único_de_identificación": "Nro. Parte / Código",
        "nro_parte_o_cdigo_nico_de_identificacin":   "Nro. Parte / Código",
        "descripción_fichaproducto":  "Descripción",
        "descripcin_fichaproducto":   "Descripción",
        "marca":                      "Marca",
        "categoría":                  "Categoría",
        "categora":                   "Categoría",
        "estado_ficha_producto":      "Estado",
        "acuerdo_marco":              "Acuerdo Marco",
        "catálogo":                   "Catálogo",
        "catalogo":                   "Catálogo",
        "precio_referencia":          "Precio Ref. (S/)",
        "precio_min":                 "Precio Mín. (S/)",
        "precio_max":                 "Precio Máx. (S/)",
        "precio_mediana":             "Precio Mediana (S/)",
        "precio_volatilidad":         "Volatilidad (%)",
        "n_ordenes_precio":           "Nro. Órdenes",
        "ficha_tcnica":               "Ficha Técnica (URL)",
        "ficha_técnica":              "Ficha Técnica (URL)",
        "imagen":                     "Imagen (URL)",
        "fecha_extraccion":           "Fecha Extracción",
        "fecha_primera_carga":        "Fecha Primera Carga",
        "precio_actualizado_at":      "Precios actualizados",
    }
    COL_WIDTHS = {
        "Nro. Parte / Código":  24,
        "Descripción":          52,
        "Marca":                22,
        "Categoría":            32,
        "Estado":               14,
        "Acuerdo Marco":        28,
        "Catálogo":             20,
        "Precio Ref. (S/)":     17,
        "Precio Mín. (S/)":     17,
        "Precio Máx. (S/)":     17,
        "Precio Mediana (S/)":  17,
        "Volatilidad (%)":      16,
        "Nro. Órdenes":         14,
        "Ficha Técnica (URL)":  14,
        "Imagen (URL)":         14,
    }
    MONEY_COLS = {"Precio Ref. (S/)", "Precio Mín. (S/)", "Precio Máx. (S/)", "Precio Mediana (S/)"}
    PCT_COLS   = {"Volatilidad (%)"}

    display_cols = [COL_LABELS.get(c, c.replace("_", " ").title()) for c in cols]

    # ── Styles ─────────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1E3A8A")
    TITLE_FILL = PatternFill("solid", fgColor="0F172A")
    ALT_FILL   = PatternFill("solid", fgColor="EFF6FF")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    BOLD       = Font(bold=True, name="Calibri")
    BASE_FONT  = Font(name="Calibri", size=9)
    thin       = Side(style="thin", color="CBD5E1")
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    WRAP       = Alignment(wrap_text=True, vertical="top")
    TOP        = Alignment(vertical="top")
    MONEY_FMT  = '#,##0.00'
    PCT_FMT    = '0.00"%"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Fichas Producto"

    # ── Title row ─────────────────────────────────────────────────────────
    active_filters_info = []
    if marca:        active_filters_info.append(f"Marca: {marca}")
    if categoria:    active_filters_info.append(f"Categoría: {categoria}")
    if estado:       active_filters_info.append(f"Estado: {estado}")
    if search:       active_filters_info.append(f"Búsqueda: {search}")
    if acuerdo_marco:active_filters_info.append(f"Acuerdo Marco: {acuerdo_marco}")
    if con_precio:   active_filters_info.append("Solo con precio")

    title_txt = "Fichas Producto — CEAM AUDITOR"
    if active_filters_info:
        title_txt += f"  |  Filtros: {' · '.join(active_filters_info)}"
    title_txt += f"  |  {len(rows):,} registros  |  Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    ncols = len(cols)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    tc = ws["A1"]
    tc.value = title_txt
    tc.fill = TITLE_FILL
    tc.font = TITLE_FONT
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # ── Header row ────────────────────────────────────────────────────────
    for ci, label in enumerate(display_cols, start=1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(label, 18)
    ws.row_dimensions[2].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────
    import decimal as _decimal
    def _safe_val(v):
        """Normalize PostgreSQL types to values openpyxl can handle."""
        if v is None:
            return None
        if isinstance(v, _decimal.Decimal):
            return float(v)
        # strip timezone from datetime so openpyxl doesn't choke
        if hasattr(v, 'tzinfo') and hasattr(v, 'replace') and v.tzinfo is not None:
            return v.replace(tzinfo=None)
        return v

    for ri, row in enumerate(rows, start=3):
        is_alt = ri % 2 == 0
        for ci, (val, label) in enumerate(zip(row, display_cols), start=1):
            cell = ws.cell(row=ri, column=ci, value=_safe_val(val))
            if is_alt:
                cell.fill = ALT_FILL
            cell.border = BORDER
            cell.font = BASE_FONT
            if label in MONEY_COLS:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif label in PCT_COLS:
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif label in ("Descripción", "Acuerdo Marco"):
                cell.alignment = WRAP
            else:
                cell.alignment = TOP

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"

    try:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {e}")

    parts = [x[:20].replace(" ", "_") for x in [marca, categoria] if x]
    slug = "_".join(parts) if parts else "todas"
    fname = f"fichas_{slug}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/catalog")
def get_catalog_api(
    marca: Optional[str] = Query(None, description="Filter by brand, e.g. 'KENYA TECHNOLOGY'"),
    categoria: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=2000),
    db: Session = Depends(get_db),
):
    """
    Product catalog API — returns fichas as a clean, normalized JSON list.

    - Results are **deduplicated by nro_parte** (latest extraction wins).
    - All spec columns are mapped to standard snake_case keys.
    - Designed to be consumed by external agents and product websites.

    Example:
        GET /api/v1/fichas/catalog?marca=KENYA+TECHNOLOGY&limit=500
    """
    import unicodedata
    import re as _re

    cols = _safe_col(db)
    if not cols:
        return {"marca": marca, "total": 0, "page": {"skip": skip, "limit": limit}, "items": []}

    col_set = set(cols)

    # ── Column finder (exact → substring) ────────────────────────────────
    def find_col(*candidates) -> Optional[str]:
        for c in candidates:
            if c in col_set:
                return c
        for c in candidates:
            m = next((col for col in col_set if c.lower() in col.lower()), None)
            if m:
                return m
        return None

    # ── Identify key columns ─────────────────────────────────────────────
    NRO_PARTE_COL = find_col(
        "nro_parte_o_código_único_de_identificación",
        "nro_parte_o_cdigo_nico_de_identificacin",
        "nro_parte", "codigo", "código",
    )
    MODELO_COL = find_col("modelo", "denominación_comercial", "denominacion_comercial", "nombre_comercial")
    DESC_COL   = find_col(
        "descripción_fichaproducto", "descripcin_fichaproducto", "descripcion", "descripción",
    )
    CAT_COL    = find_col("categoría", "categora", "categoria")
    MARCA_COL  = find_col("marca")
    ESTADO_COL = find_col("estado_ficha_producto", "estado")
    IMG_COL    = find_col("imagen", "imagen_url", "imagen_producto", "foto")
    PDF_COL    = find_col("ficha_técnica", "ficha_tcnica", "ficha_tecnica")
    MODEL_SRC  = MODELO_COL or DESC_COL

    # ── Spec key patterns (pattern_substring → clean_key) ─────────────────
    SPEC_PATTERNS: dict = {
        "procesador":        ["procesador", "cpu", "tipo_de_procesador"],
        "ram":               ["ram", "memoria_ram", "memoria"],
        "almacenamiento":    ["almacenamiento", "disco_duro", "ssd", "hdd", "nvme",
                              "unidad_de_almacenamiento", "unidad_de_estado"],
        "lan":               ["lan", "ethernet", "red_alambrica", "red_alámbrica"],
        "wlan":              ["wlan", "wifi", "wi-fi", "wireless", "inalámbrico", "inalambrico"],
        "usb":               ["usb", "puertos_usb"],
        "vga":               ["vga", "salida_vga", "puerto_vga"],
        "hdmi":              ["hdmi", "salida_hdmi", "puerto_hdmi"],
        "sistema_operativo": ["sistema_operativo", "s.o.", "sistema_op"],
        "unidad_optica":     ["unidad_óptica", "unidad_optica", "dvd", "grabadora",
                              "unidad_de_disco"],
        "teclado":           ["teclado"],
        "mouse":             ["mouse", "ratón", "raton"],
        "suite_ofimatica":   ["suite_ofimática", "suite_ofimatica", "ofimatica", "suite",
                              "microsoft_office", "libreoffice"],
        "garantia_fabrica":  ["garantía", "garantia", "garantía_de_fábrica", "garantia_de_fabrica",
                              "garantía_del_fabricante", "garantia_del_fabricante",
                              "tiempo_de_garantía", "tiempo_de_garantia"],
        "pantalla":          ["pantalla", "display", "monitor", "tamaño_de_pantalla",
                              "tamaño_pantalla", "resolución_de_pantalla"],
    }

    col_to_spec: dict = {}
    for skey, patterns in SPEC_PATTERNS.items():
        for p in patterns:
            col_to_spec[p.lower()] = skey

    META_COLS = {c for c in [
        NRO_PARTE_COL, MODELO_COL, DESC_COL, CAT_COL, MARCA_COL, ESTADO_COL,
        IMG_COL, PDF_COL, "fecha_extraccion", "fecha_primera_carga",
        "acuerdo_marco", "catálogo", "catalogo",
        "precio_referencia", "precio_min", "precio_max", "precio_mediana",
        "precio_volatilidad", "n_ordenes_precio", "precio_actualizado_at",
    ] if c}

    # ── WHERE clause ──────────────────────────────────────────────────────
    filters: list = []
    params: dict = {}

    if MARCA_COL and marca:
        filters.append(f'"{MARCA_COL}" ILIKE :marca')
        params["marca"] = f"%{marca}%"
    if CAT_COL and categoria:
        filters.append(f'"{CAT_COL}" ILIKE :categoria')
        params["categoria"] = f"%{categoria}%"
    if ESTADO_COL and estado:
        filters.append(f'"{ESTADO_COL}" ILIKE :estado')
        params["estado"] = f"%{estado}%"
    if search:
        sc = [f'"{c}" ILIKE :search' for c in [NRO_PARTE_COL, MODEL_SRC] if c]
        if sc:
            filters.append("(" + " OR ".join(sc) + ")")
            params["search"] = f"%{search}%"

    where = ("WHERE " + " AND ".join(filters)) if filters else ""

    # ── Total count (distinct nro_parte) ──────────────────────────────────
    try:
        if NRO_PARTE_COL:
            total_count = int(
                db.execute(
                    text(f'SELECT COUNT(DISTINCT "{NRO_PARTE_COL}") FROM {_TABLE} {where}'),
                    params,
                ).scalar() or 0
            )
        else:
            total_count = int(
                db.execute(text(f"SELECT COUNT(*) FROM {_TABLE} {where}"), params).scalar() or 0
            )
    except Exception:
        total_count = 0

    # ── Query — deduplicated: one row per nro_parte (latest extraction) ───
    quoted = ", ".join(f'"{c}"' for c in cols)
    has_date = "fecha_extraccion" in col_set

    if NRO_PARTE_COL and has_date:
        sql = text(
            f'SELECT DISTINCT ON ("{NRO_PARTE_COL}") {quoted} '
            f'FROM {_TABLE} {where} '
            f'ORDER BY "{NRO_PARTE_COL}", fecha_extraccion DESC NULLS LAST '
            f'LIMIT :limit OFFSET :skip'
        )
    else:
        sql = text(
            f"SELECT {quoted} FROM {_TABLE} {where} "
            f"ORDER BY fecha_extraccion DESC NULLS LAST "
            f"LIMIT :limit OFFSET :skip"
        )

    params["limit"] = limit
    params["skip"] = skip

    try:
        rows = db.execute(sql, params).fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Query error: {e}")

    # ── Helpers ───────────────────────────────────────────────────────────
    def normalize_key(col: str) -> str:
        k = _re.sub(r'\s*\([^)]*\)', '', col)
        k = _re.sub(r'[^\w]+', '_', k, flags=_re.UNICODE)
        k = ''.join(
            c for c in unicodedata.normalize('NFD', k.strip('_').lower())
            if unicodedata.category(c) != 'Mn'
        )
        return k or col.lower()

    def resolve_spec_key(col: str) -> str:
        cl = col.lower()
        if cl in col_to_spec:
            return col_to_spec[cl]
        for pattern, skey in col_to_spec.items():
            if pattern in cl:
                return skey
        return normalize_key(col)

    def clean_text(v) -> Optional[str]:
        if v is None:
            return None
        s = str(v).strip()
        return s.upper() if s else None

    def clean_url(v) -> Optional[str]:
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    # Pre-compute spec key per column (first match wins — no duplicates)
    col_spec_keys: dict = {}
    seen_spec_keys: set = set()
    for col in cols:
        if col in META_COLS:
            continue
        key = resolve_spec_key(col)
        if key not in seen_spec_keys:
            col_spec_keys[col] = key
            seen_spec_keys.add(key)

    # ── Build response items ──────────────────────────────────────────────
    items = []
    for row in rows:
        d = dict(zip(cols, row))
        specs = {skey: clean_text(d.get(col)) for col, skey in col_spec_keys.items()}
        items.append({
            "nro_parte":         clean_text(d.get(NRO_PARTE_COL)) if NRO_PARTE_COL else None,
            "modelo":            clean_text(d.get(MODEL_SRC))      if MODEL_SRC     else None,
            "categoria":         clean_text(d.get(CAT_COL))        if CAT_COL       else None,
            "estado":            clean_text(d.get(ESTADO_COL))     if ESTADO_COL    else None,
            "imagen_url":        clean_url(d.get(IMG_COL))         if IMG_COL       else None,
            "ficha_tecnica_url": clean_url(d.get(PDF_COL))         if PDF_COL       else None,
            "specs":             specs,
        })

    return {
        "marca":  marca,
        "total":  total_count,
        "page":   {"skip": skip, "limit": limit},
        "items":  items,
    }


# ---------------------------------------------------------------------------
# GET /fichas/video-specs
# ---------------------------------------------------------------------------
@router.get("/video-specs")
def fichas_video_specs():
    """
    Devuelve los specs de video/gráficos extraídos de las fichas técnicas PDF
    de Kenya Technology.

    Los datos son generados por scripts/extract_video_specs.py y almacenados
    en backend/app/data/kenya_video_specs.json.
    """
    import json
    from pathlib import Path

    data_file = Path(__file__).parent.parent.parent / "data" / "kenya_video_specs.json"

    if not data_file.exists():
        return {
            "detail": "No hay datos. Ejecuta el script scripts/extract_video_specs.py primero.",
            "marca": "KENYA TECHNOLOGY",
            "total": 0,
            "items": [],
        }

    try:
        items = json.loads(data_file.read_text(encoding="utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error leyendo datos: {exc}")

    return {
        "marca": "KENYA TECHNOLOGY",
        "total": len(items),
        "items": items,
    }


@router.post("/video-specs/refresh", status_code=202)
def fichas_video_specs_refresh():
    """
    Dispara la tarea Celery de extracción de specs de video en background.
    El resultado queda disponible en GET /fichas/video-specs al terminar (~15 min).
    """
    from app.worker.tasks import refresh_video_specs_task

    task = refresh_video_specs_task.delay()
    return {
        "task_id": task.id,
        "status": "queued",
        "message": "Extracción iniciada. Consulta GET /fichas/video-specs cuando termine.",
    }


@router.get("/summary")
def fichas_summary(
    acuerdo_marco: Optional[str] = Query(None),
    catalogo: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    con_precio: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
):
    """Filtered aggregate stats for KPI cards (total, con_precio, sin_precio, volatility)."""
    cols = _safe_col(db)
    col_set = set(cols)
    _, params, where_clause = _build_fichas_where(
        col_set, acuerdo_marco, catalogo, categoria, marca, estado, search, con_precio
    )
    has_ref = "precio_referencia" in col_set
    has_vol = "precio_volatilidad" in col_set
    try:
        if has_ref and has_vol:
            row = db.execute(text(
                f"SELECT COUNT(*) as total, "
                f"COUNT(precio_referencia) as con_precio, "
                f"COUNT(*) FILTER (WHERE precio_volatilidad < 20) as vol_baja, "
                f"COUNT(*) FILTER (WHERE precio_volatilidad BETWEEN 20 AND 50) as vol_media, "
                f"COUNT(*) FILTER (WHERE precio_volatilidad > 50) as vol_alta "
                f"FROM {_TABLE} {where_clause}"
            ), params).fetchone()
            total, con_p, v_baja, v_media, v_alta = [x or 0 for x in row]
        elif has_ref:
            row = db.execute(text(
                f"SELECT COUNT(*), COUNT(precio_referencia) FROM {_TABLE} {where_clause}"
            ), params).fetchone()
            total, con_p = row[0] or 0, row[1] or 0
            v_baja = v_media = v_alta = 0
        else:
            total = db.execute(text(f"SELECT COUNT(*) FROM {_TABLE} {where_clause}"), params).scalar() or 0
            con_p = v_baja = v_media = v_alta = 0
        return {
            "total": total,
            "con_precio": con_p,
            "sin_precio": total - con_p,
            "coverage_pct": round(con_p / total * 100, 1) if total > 0 else 0.0,
            "volatilidad": {"baja": v_baja, "media": v_media, "alta": v_alta},
        }
    except Exception:
        return {
            "total": 0, "con_precio": 0, "sin_precio": 0, "coverage_pct": 0.0,
            "volatilidad": {"baja": 0, "media": 0, "alta": 0},
        }


@router.get("/filters/{col_name}")
def get_fichas_filter_values(col_name: str, db: Session = Depends(get_db)):
    """Return distinct non-null values for a fichas column (for Excel-like dropdowns)."""
    allowed = {"marca", "categoría", "categora", "acuerdo_marco", "estado_ficha_producto"}
    cols = _safe_col(db)
    col_set = set(cols)

    # Accept either spelling; find actual column name in DB
    candidates = [col_name]
    if col_name == "categoria":
        candidates = ["categoría", "categora"]
    matched = next((c for c in candidates if c in col_set), None)
    if matched is None or matched not in {c for c in col_set if c in allowed or col_name in allowed}:
        # Still allow if the matched col exists even if alias differs
        if matched is None:
            return {"values": []}

    try:
        rows = db.execute(
            text(f'SELECT DISTINCT "{matched}" FROM {_TABLE} WHERE "{matched}" IS NOT NULL AND "{matched}" != \'\' ORDER BY "{matched}"')
        ).fetchall()
        return {"values": [r[0] for r in rows if r[0]]}
    except Exception:
        return {"values": []}


@router.get("/stats")
def get_fichas_stats(db: Session = Depends(get_db)):
    """Aggregated statistics for the fichas dashboard panel."""
    cols = _safe_col(db)
    if not cols:
        return {
            "total_fichas": 0,
            "by_acuerdo": [],
            "by_categoria": [],
            "by_estado": [],
            "by_marca": [],
        }

    col_set = set(cols)

    def _agg(col: str, label: str):
        if col not in col_set:
            return []
        try:
            rows = db.execute(
                text(
                    f'SELECT "{col}", COUNT(*) as total FROM {_TABLE}'
                    f' GROUP BY "{col}" ORDER BY total DESC LIMIT 20'
                )
            ).fetchall()
            return [{"name": r[0] or "S/D", "total": r[1]} for r in rows]
        except Exception:
            return []

    total = 0
    try:
        total = db.execute(text(f"SELECT COUNT(*) FROM {_TABLE}")).scalar() or 0
    except Exception:
        pass

    return {
        "total_fichas": total,
        "by_acuerdo": _agg("acuerdo_marco", "acuerdo"),
        "by_categoria": _agg("categoría" if "categoría" in col_set else "categora", "categoria"),
        "by_estado": _agg("estado_ficha_producto", "estado"),
        "by_marca": _agg("marca", "marca"),
    }


@router.get("/precio-stats")
def get_precio_stats(db: Session = Depends(get_db)):
    """Coverage and volatility stats for enriched prices."""
    cols = _safe_col(db)
    if "precio_referencia" not in cols:
        return {"total": 0, "con_precio": 0, "sin_precio": 0, "coverage_pct": 0.0, "enriquecido_at": None, "volatilidad": {"baja": 0, "media": 0, "alta": 0}}
    try:
        total = db.execute(text(f"SELECT COUNT(*) FROM {_TABLE}")).scalar() or 0
        con_precio = db.execute(text(f"SELECT COUNT(*) FROM {_TABLE} WHERE precio_referencia IS NOT NULL")).scalar() or 0
        last_upd = db.execute(text(f"SELECT MAX(precio_actualizado_at) FROM {_TABLE}")).scalar()
        vol = db.execute(text(
            f"SELECT "
            f"COUNT(*) FILTER (WHERE precio_volatilidad < 20) as baja, "
            f"COUNT(*) FILTER (WHERE precio_volatilidad BETWEEN 20 AND 50) as media, "
            f"COUNT(*) FILTER (WHERE precio_volatilidad > 50) as alta "
            f"FROM {_TABLE} WHERE precio_referencia IS NOT NULL"
        )).fetchone()
        return {
            "total": total,
            "con_precio": con_precio,
            "sin_precio": total - con_precio,
            "coverage_pct": round(con_precio / total * 100, 1) if total > 0 else 0.0,
            "enriquecido_at": str(last_upd) if last_upd else None,
            "volatilidad": {"baja": vol[0] or 0, "media": vol[1] or 0, "alta": vol[2] or 0},
        }
    except Exception:
        return {"total": 0, "con_precio": 0, "sin_precio": 0, "coverage_pct": 0.0, "enriquecido_at": None, "volatilidad": {"baja": 0, "media": 0, "alta": 0}}


@router.post("/enrich-precios")
def enrich_precios(db: Session = Depends(get_db)):
    """
    Match fichas_producto ↔ purchase_orders by nro_parte and compute a
    canonical reference price using epsilon-neighborhood mode clustering.

    Algorithm (Discrete Math — proximity equivalence classes):
      - Sort all prices for a given nro_parte.
      - Build clusters: greedy scan; a price joins the current cluster if it
        falls within ε=5% of the cluster’s first element (anchor).
        Formally: (p₁, p₂) ∈ R  ⇔  p₂ ≤ p₁ × (1+ε).
      - Select the densest cluster (most orders at that price zone).
      - Canonical price = median of the densest cluster.
      - Volatility = (max − min) / global_median × 100  (% spread).
    """
    # 1. Add price columns to fichas_producto if not present
    price_cols = [
        ("precio_referencia", "NUMERIC(14,4)"),
        ("precio_min", "NUMERIC(14,4)"),
        ("precio_max", "NUMERIC(14,4)"),
        ("precio_mediana", "NUMERIC(14,4)"),
        ("precio_volatilidad", "NUMERIC(7,2)"),
        ("n_ordenes_precio", "INTEGER"),
        ("precio_actualizado_at", "TIMESTAMP WITH TIME ZONE"),
    ]
    try:
        for col_name, col_type in price_cols:
            db.execute(text(f"ALTER TABLE {_TABLE} ADD COLUMN IF NOT EXISTS {col_name} {col_type}"))
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al agregar columnas de precio: {e}")

    # 2. Detect nro_parte key column in fichas_producto
    cols = _safe_col(db)
    nro_col = next((c for c in cols if c.startswith("nro_parte")), None)
    if not nro_col:
        raise HTTPException(status_code=422, detail="No se encontró columna nro_parte en fichas_producto")

    # 3. Gather all prices from purchase_orders grouped by nro_parte
    # purchase_orders.nro_parte stores a JSON array: [{"nro_parte": "...", "precio_unitario": N, ...}]
    # We expand that JSON with jsonb_array_elements so we get one row per product.
    # Normalize keys: UPPER + STRIP to avoid case-sensitivity mismatches.
    raw = db.execute(text(
        """
        SELECT
            elem->>'nro_parte'                          AS nro_parte,
            (elem->>'precio_unitario')::numeric          AS precio_unitario
        FROM purchase_orders
        CROSS JOIN LATERAL jsonb_array_elements(
            CASE
                WHEN nro_parte IS NOT NULL
                     AND nro_parte NOT IN ('', 'null', '[]')
                     AND nro_parte LIKE '[%'
                THEN nro_parte::jsonb
                ELSE '[]'::jsonb
            END
        ) AS elem
        WHERE (elem->>'precio_unitario')::numeric > 0
          AND elem->>'nro_parte' IS NOT NULL
          AND elem->>'nro_parte' <> ''
        """
    )).fetchall()
    price_map: dict = defaultdict(list)
    for nro, precio in raw:
        normalized = str(nro).strip().upper()
        if normalized:  # skip empty strings after normalization
            price_map[normalized].append(float(precio))

    # Diagnostic: log first 5 keys so mismatches can be spotted quickly
    sample_po_keys = list(price_map.keys())[:5]
    import logging as _log
    _log.getLogger("ceam.enrich").info(
        "price_map built: %d unique nro_parte from purchase_orders. Sample: %s",
        len(price_map), sample_po_keys
    )

    # 4. Epsilon-neighborhood mode clustering
    def canonical_price(precios: list) -> dict:
        precios_s = sorted(precios)
        EPS = 0.05  # 5% proximity tolerance
        clusters, current = [], [precios_s[0]]
        for p in precios_s[1:]:
            if current[0] > 0 and p <= current[0] * (1 + EPS):
                current.append(p)
            else:
                clusters.append(current)
                current = [p]
        clusters.append(current)
        best = max(clusters, key=len)
        canonical = statistics.median(best)
        med_g = statistics.median(precios_s)
        volatilidad = round((max(precios_s) - min(precios_s)) / med_g * 100, 2) if med_g > 0 else 0.0
        return {
            "precio_referencia": round(canonical, 4),
            "precio_min": round(min(precios_s), 4),
            "precio_max": round(max(precios_s), 4),
            "precio_mediana": round(med_g, 4),
            "precio_volatilidad": volatilidad,
            "n_ordenes_precio": len(precios_s),
        }

    # 5. Update fichas_producto
    now = datetime.now(tz=timezone.utc)
    fichas_keys = db.execute(text(f'SELECT "{nro_col}" FROM {_TABLE} WHERE "{nro_col}" IS NOT NULL')).fetchall()

    enriched = 0
    not_found = 0
    for (nro_val,) in fichas_keys:
        original_key = str(nro_val).strip()   # preserve original case for WHERE
        key = original_key.upper()             # normalized key for price_map lookup
        if not key or key in ("", "NAN", "NONE"):
            not_found += 1
            continue
        prices = price_map.get(key)
        if not prices:
            not_found += 1
            continue
        cp = canonical_price(prices)
        db.execute(text(
            f'UPDATE {_TABLE} SET '
            f'precio_referencia = :pr, precio_min = :pmin, precio_max = :pmax, '
            f'precio_mediana = :pmed, precio_volatilidad = :pvol, '
            f'n_ordenes_precio = :n, precio_actualizado_at = :ts '
            f'WHERE "{nro_col}" = :key'
        ), {"pr": cp["precio_referencia"], "pmin": cp["precio_min"], "pmax": cp["precio_max"],
            "pmed": cp["precio_mediana"], "pvol": cp["precio_volatilidad"],
            "n": cp["n_ordenes_precio"], "ts": now, "key": original_key})
        enriched += 1

    db.commit()
    total = len(fichas_keys)
    return {
        "enriched": enriched,
        "not_found": not_found,
        "total_fichas": total,
        "coverage_pct": round(enriched / total * 100, 1) if total > 0 else 0.0,
        "_debug": {
            "price_map_size": len(price_map),
            "sample_po_keys": list(price_map.keys())[:5],
        },
    }


@router.delete("/all")
def delete_all_fichas(db: Session = Depends(get_db)):
    """Truncate the fichas_producto table. Used to clear duplicates before a clean re-scrape."""
    try:
        db.execute(text(f"TRUNCATE TABLE {_TABLE} RESTART IDENTITY"))
        db.commit()
        return {"deleted": True, "message": "Tabla fichas_producto vaciada correctamente."}
    except Exception as e:
        db.rollback()
        raise


@router.get("/alertas-suspendidas")
async def get_alertas_suspendidas(
    acuerdo_marco: str = Query("EXT-CE-2022-5", description="Código del Acuerdo Marco a escanear"),
    db: Session = Depends(get_db)
):
    """
    Endpoint para n8n. Ejecuta el scraper de Módulo 2 en vivo y
    devuelve las fichas que pasaron de 'Ofertada' → 'Suspendida'.
    """
    from app.services.fichas_scraper import run_module_2, AGREEMENT_SELECTOR

    # Build the CSS selector from the agreement code
    selector = f'div[data-agreement*="{acuerdo_marco}"]'
            
    try:
        from app.db.database import engine as app_engine
        result = await run_module_2(
            engine=app_engine,
            agreement_selector=selector,
            agreement_code=acuerdo_marco,
            cleanup=True
        )
        
        deltas = result.get("deltas_suspendidas", [])
        
        # Agrupar por marca
        datos = {}
        for d in deltas:
            marca = d["marca"].upper() if d["marca"] else "OTRAS"
            if marca not in datos:
                datos[marca] = []
            datos[marca].append({
                "nro_parte": d["nro_parte"],
                "descripcion": d["descripcion"],
                "anterior": d["anterior"],
                "actual": d["actual"]
            })
            
        resumen = {}
        # Para cada marca que tuvo caídas, consultar en base de datos cuántas 'Ofertadas' le quedan
        cols = _safe_col(db)
        if "marca" in cols and "estado_ficha_producto" in cols:
            for marca in datos.keys():
                count = db.execute(
                    text(
                        f"SELECT COUNT(*) FROM {_TABLE} "
                        f"WHERE UPPER(marca) = :m AND UPPER(estado_ficha_producto) LIKE '%OFERTADA%'"
                    ),
                    {"m": marca}
                ).scalar() or 0
                resumen[marca] = {"ofertadas_actuales": count}
                
        return {
            "hayAlertas": len(deltas) > 0,
            "total_suspendidas": len(deltas),
            "datos": datos,
            "resumen": resumen,
            "meta": {
                "filepath": result.get("filepath"),
                "rows_processed": result.get("rows_processed"),
                "inserted": result.get("inserted"),
                "updated": result.get("updated")
            }
        }
        
    except Exception as e:
        import traceback
        return {
            "error": True,
            "message": str(e),
            "trace": traceback.format_exc()
        }
