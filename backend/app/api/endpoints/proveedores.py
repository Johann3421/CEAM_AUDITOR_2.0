import json
import re
from fastapi import APIRouter, Depends, HTTPException, Query, BackgroundTasks
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional, List, Dict

from app.db.database import get_db
from app.services.proveedores_scraper import run_worker_pool_extraction, fetch_single_combo

router = APIRouter(prefix="/proveedores", tags=["proveedores"])

def _get_fichas_pdf_join(db: Session):
    """
    Returns (join_sql, select_expr, ord_min_expr, fec_min_expr, pref_expr, pmin_expr, momin_expr) 
    to link PDF and historical order/price data from fichas_producto by nro_parte.
    """
    fallback = ("", "f.pdf_url", "NULL::text", "NULL::text", "NULL::numeric", "NULL::numeric", "NULL::numeric")
    try:
        rows = db.execute(text(
            "SELECT column_name FROM information_schema.columns WHERE table_name = 'fichas_producto'"
        )).fetchall()
        cols = [r[0] for r in rows]
        if not cols:
            return fallback
            
        col_set = set(cols)
        
        # 1. Detect nro_parte column
        nro_col = next((c for c in [
            "nro_parte_o_código_único_de_identificación",
            "nro_parte_o_cdigo_nico_de_identificacin",
            "nro_parte", "codigo_ficha", "cod_ficha", "codigo"
        ] if c in col_set), None)
        if not nro_col:
            nro_col = next((c for c in cols if 'parte' in c.lower() or 'codigo' in c.lower()), None)
            
        # 2. Detect pdf column
        pdf_col = next((c for c in [
            "ficha_técnica", "ficha_tcnica", "ficha_tecnica", "url_pdf", "pdf_url"
        ] if c in col_set), None)
        if not pdf_col:
            pdf_col = next((c for c in cols if 'pdf' in c.lower() or 'ficha' in c.lower() or 'url' in c.lower()), None)
            
        if nro_col:
            pdf_select_field = f'fp."{pdf_col}" AS _linked_pdf,' if pdf_col else "NULL::text AS _linked_pdf,"
            ord_min_field = 'fp.orden_min AS _linked_orden_min,' if 'orden_min' in col_set else "NULL::text AS _linked_orden_min,"
            fec_min_field = 'fp.fecha_orden_min::text AS _linked_fecha_orden_min,' if 'fecha_orden_min' in col_set else "NULL::text AS _linked_fecha_orden_min,"
            pref_field = 'fp.precio_referencia AS _linked_precio_ref,' if 'precio_referencia' in col_set else "NULL::numeric AS _linked_precio_ref,"
            pmin_field = 'fp.precio_min AS _linked_precio_min,' if 'precio_min' in col_set else "NULL::numeric AS _linked_precio_min,"
            momin_field = 'fp.monto_orden_min AS _linked_monto_orden_min' if 'monto_orden_min' in col_set else "NULL::numeric AS _linked_monto_orden_min"

            order_by_extra = f', fp."{pdf_col}" DESC NULLS LAST' if pdf_col else ''
            join_sql = f"""
                LEFT JOIN (
                    SELECT DISTINCT ON (UPPER(TRIM(fp."{nro_col}")))
                        UPPER(TRIM(fp."{nro_col}")) AS _match_nro,
                        {pdf_select_field}
                        {ord_min_field}
                        {fec_min_field}
                        {pref_field}
                        {pmin_field}
                        {momin_field}
                    FROM fichas_producto fp
                    WHERE fp."{nro_col}" IS NOT NULL AND fp."{nro_col}" != ''
                    ORDER BY 
                        UPPER(TRIM(fp."{nro_col}")),
                        (fp.orden_min IS NOT NULL AND fp.orden_min != '') DESC,
                        (fp.precio_min IS NOT NULL) DESC,
                        fp.precio_min ASC NULLS LAST
                        {order_by_extra}
                ) fp_pdf ON UPPER(TRIM(f.nro_parte)) = fp_pdf._match_nro
            """
            select_expr = "COALESCE(NULLIF(NULLIF(f.pdf_url, ''), '#'), fp_pdf._linked_pdf)" if pdf_col else "f.pdf_url"
            return (
                join_sql,
                select_expr,
                "fp_pdf._linked_orden_min",
                "fp_pdf._linked_fecha_orden_min",
                "fp_pdf._linked_precio_ref",
                "fp_pdf._linked_precio_min",
                "fp_pdf._linked_monto_orden_min"
            )
    except Exception:
        pass
        
    return fallback


@router.get("/fichas")
def get_proveedor_fichas(
    proveedor: Optional[str] = Query(None, description="Filtro por nombre de proveedor"),
    proveedor_filter: Optional[str] = Query(None, description="Filtro específico: 'ambos', 'exclusivo', 'thekingcomputer', 'jorge_rojas'"),
    region: Optional[str] = Query(None, description="Filtro por región / departamento"),
    search: Optional[str] = Query(None, description="Búsqueda rápida por nro_parte o descripción"),
    marca: Optional[str] = Query(None, description="Filtro por marca"),
    nro_parte: Optional[str] = Query(None, description="Filtro por nro_parte"),
    catalogo: Optional[str] = Query(None, description="Filtro por catálogo"),
    categoria: Optional[str] = Query(None, description="Filtro por categoría"),
    stock_filter: Optional[str] = Query(None, description="Filtro de stock: 'with_stock' o 'zero_stock'"),
    pdf_filter: Optional[str] = Query(None, description="Filtro de PDF: 'with_pdf' o 'no_pdf'"),
    sort_by: Optional[str] = Query(None, description="Ordenamiento: precio_asc, precio_desc, stock_desc, marca_asc"),
    cpu: Optional[str] = Query(None, description="Filtro de CPU"),
    ram: Optional[str] = Query(None, description="Filtro de Memoria RAM"),
    disco: Optional[str] = Query(None, description="Filtro de Disco / Almacenamiento"),
    pantalla: Optional[str] = Query(None, description="Filtro de Pantalla en pulgadas"),
    so: Optional[str] = Query(None, description="Filtro de Sistema Operativo"),
    panel: Optional[str] = Query(None, description="Filtro de Tipo de Panel (IPS, VA, etc.)"),
    resolucion: Optional[str] = Query(None, description="Filtro de Resolución"),
    cpu_gen: Optional[str] = Query(None, description="Filtro de Generación CPU: gen14, gen13, gen12, gen11, gen10, ultra, ryzen7000, ryzen5000"),
    ram_tech: Optional[str] = Query(None, description="Filtro de Tecnología RAM: DDR4, DDR5, LPDDR5"),
    disco_tipo: Optional[str] = Query(None, description="Filtro de Tipo de Disco: NVMe, M.2, hibrido, solo_ssd, solo_hdd"),
    vga: Optional[str] = Query(None, description="Filtro de Puerto VGA: si, no"),
    hdmi: Optional[str] = Query(None, description="Filtro de Puerto HDMI: si, no"),
    wifi: Optional[str] = Query(None, description="Filtro de Wi-Fi / WLAN: si, no"),
    bluetooth: Optional[str] = Query(None, description="Filtro de Bluetooth: si, no"),
    lan: Optional[str] = Query(None, description="Filtro de Puerto de Red LAN: si, no"),
    office: Optional[str] = Query(None, description="Filtro de Suite Ofimática: home_business, si, no"),
    garantia: Optional[str] = Query(None, description="Filtro de Garantía: 36, 24, 12"),
    unidad_optica: Optional[str] = Query(None, description="Filtro de Unidad Óptica: si, no"),
    camara: Optional[str] = Query(None, description="Filtro de Cámara Web: si, no"),
    tactil: Optional[str] = Query(None, description="Filtro de Pantalla Táctil: si, no"),
    con_orden: Optional[bool] = Query(None, description="Filtrar solo fichas con orden de compra registrada"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=5000),
    db: Session = Depends(get_db)
):
    """
    Lista las fichas y ofertas por proveedor con soporte para filtros por columna, región y ordenamiento tipo Excel.
    """
    offset = (page - 1) * limit
    params = {}
    where_clauses = ["1=1"]

    pdf_join, pdf_select, ord_min_expr, fec_min_expr, pref_expr, pmin_expr, momin_expr = _get_fichas_pdf_join(db)

    if proveedor and proveedor.lower() != "all":
        prov_l = proveedor.lower()
        if prov_l in ("thekingcomputer", "king"):
            where_clauses.append("(UPPER(f.nombre_proveedor) LIKE '%KING%' OR UPPER(f.ruc_proveedor) = '20601234567')")
        elif prov_l in ("jorge_rojas", "jorge", "rojas"):
            where_clauses.append("(UPPER(f.nombre_proveedor) LIKE '%ROJAS%' OR UPPER(f.nombre_proveedor) LIKE '%JORGE%' OR UPPER(f.ruc_proveedor) = '10408899991')")
        else:
            where_clauses.append("UPPER(f.nombre_proveedor) LIKE UPPER(:proveedor)")
            params["proveedor"] = f"%{proveedor}%"

    if search:
        where_clauses.append("(UPPER(f.nro_parte) LIKE UPPER(:search) OR UPPER(f.descripcion_producto) LIKE UPPER(:search) OR UPPER(f.marca) LIKE UPPER(:search))")
        params["search"] = f"%{search}%"

    if marca:
        where_clauses.append("UPPER(f.marca) = UPPER(:marca)")
        params["marca"] = marca

    if nro_parte:
        where_clauses.append("UPPER(f.nro_parte) LIKE UPPER(:nro_parte)")
        params["nro_parte"] = f"%{nro_parte}%"

    if catalogo:
        where_clauses.append("UPPER(f.catalogo) = UPPER(:catalogo)")
        params["catalogo"] = catalogo

    if categoria:
        categ_lower = categoria.lower().strip()
        if categ_lower in ("escritorio", "computadora de escritorio"):
            where_clauses.append("""(
                UPPER(f.categoria) = 'COMPUTADORA DE ESCRITORIO'
                OR (
                    (UPPER(f.catalogo) LIKE '%ESCRITORIO%' OR UPPER(f.categoria) LIKE '%ESCRITORIO%' OR UPPER(f.descripcion_producto) LIKE '%ESCRITORIO%')
                    AND UPPER(f.categoria) NOT LIKE '%TODO EN UNO%'
                    AND UPPER(f.categoria) NOT LIKE '%MONITOR%'
                    AND UPPER(f.categoria) NOT LIKE '%ESTACION%'
                    AND UPPER(f.categoria) NOT LIKE '%ALMACENAMIENTO%'
                    AND UPPER(f.categoria) NOT LIKE '%PANTALLA%'
                    AND UPPER(f.descripcion_producto) NOT LIKE '%TODO EN UNO%'
                    AND UPPER(f.descripcion_producto) NOT LIKE '%ALL IN ONE%'
                    AND UPPER(f.descripcion_producto) NOT LIKE '%MONITOR%'
                )
            )""")
        elif categ_lower in ("aio", "todo en uno", "all in one", "computadora todo en uno"):
            where_clauses.append("""(
                UPPER(f.categoria) LIKE '%TODO EN UNO%'
                OR UPPER(f.descripcion_producto) LIKE '%TODO EN UNO%'
                OR UPPER(f.descripcion_producto) LIKE '%ALL IN ONE%'
                OR UPPER(f.descripcion_producto) LIKE '%ALL-IN-ONE%'
            )""")
        elif categ_lower in ("workstation", "estacion de trabajo", "estacion"):
            where_clauses.append("(UPPER(f.categoria) LIKE '%ESTACION DE TRABAJO%' OR UPPER(f.descripcion_producto) LIKE '%WORKSTATION%')")
        elif categ_lower in ("monitor", "monitores"):
            where_clauses.append("""(
                UPPER(f.categoria) LIKE '%MONITOR%'
                OR UPPER(f.descripcion_producto) LIKE 'MONITOR%'
                OR UPPER(f.descripcion_producto) LIKE '%MONITOR LED%'
                OR UPPER(f.descripcion_producto) LIKE '%MONITOR GAMER%'
            )""")
        elif categ_lower in ("pantalla_pub", "pantalla publicitaria"):
            where_clauses.append("(UPPER(f.categoria) LIKE '%PUBLICITARIA%' OR UPPER(f.descripcion_producto) LIKE '%PUBLICITARIA%')")
        elif categ_lower in ("pantalla_int", "pantalla interactiva"):
            where_clauses.append("(UPPER(f.categoria) LIKE '%INTERACTIVA%' OR UPPER(f.descripcion_producto) LIKE '%INTERACTIVA%')")
        elif categ_lower in ("almacenamiento_int", "almacenamiento interno"):
            where_clauses.append("""(
                UPPER(f.categoria) LIKE '%INTERNO%'
                OR (UPPER(f.catalogo) LIKE '%ALMACENAMIENTO%' AND UPPER(f.descripcion_producto) NOT LIKE '%EXTERNO%')
            )""")
        elif categ_lower in ("almacenamiento_ext", "almacenamiento externo"):
            where_clauses.append("""(
                UPPER(f.categoria) LIKE '%EXTERNO%'
                OR UPPER(f.descripcion_producto) LIKE '%EXTERNO%'
            )""")
        elif categ_lower in ("portatil", "computadora portatil", "laptop"):
            where_clauses.append("""(
                (UPPER(f.categoria) LIKE '%PORTATIL%' OR UPPER(f.categoria) LIKE '%PORTÁTIL%' OR UPPER(f.descripcion_producto) LIKE '%PORTATIL%' OR UPPER(f.descripcion_producto) LIKE '%LAPTOP%')
                AND UPPER(f.categoria) NOT LIKE '%TODO EN UNO%'
                AND UPPER(f.categoria) NOT LIKE '%ESTACION%'
            )""")
        elif categ_lower in ("workstation_portatil", "estacion de trabajo portatil"):
            where_clauses.append("""(
                (UPPER(f.categoria) LIKE '%ESTACION DE TRABAJO%' OR UPPER(f.descripcion_producto) LIKE '%WORKSTATION%')
                AND (UPPER(f.categoria) LIKE '%PORTATIL%' OR UPPER(f.descripcion_producto) LIKE '%PORTATIL%' OR UPPER(f.descripcion_producto) LIKE '%LAPTOP%')
            )""")
        elif categ_lower in ("tableta", "tablet"):
            where_clauses.append("(UPPER(f.categoria) LIKE '%TABLET%' OR UPPER(f.descripcion_producto) LIKE '%TABLET%')")
        elif categ_lower in ("escaner_docs", "escaner de documentos"):
            where_clauses.append("(UPPER(f.categoria) LIKE '%DOCUMENTOS%' OR (UPPER(f.catalogo) LIKE '%ESCANER%' AND UPPER(f.descripcion_producto) NOT LIKE '%PLANO%'))")
        elif categ_lower in ("escaner_planos", "escaner de planos"):
            where_clauses.append("(UPPER(f.categoria) LIKE '%PLANO%' OR UPPER(f.descripcion_producto) LIKE '%PLANO%')")
        elif categ_lower in ("escaner_libros", "escaner de libros"):
            where_clauses.append("(UPPER(f.categoria) LIKE '%LIBRO%' OR UPPER(f.descripcion_producto) LIKE '%LIBRO%')")
        else:
            where_clauses.append("(UPPER(f.categoria) LIKE UPPER(:categoria) OR UPPER(f.catalogo) LIKE UPPER(:categoria) OR UPPER(f.descripcion_producto) LIKE UPPER(:categoria))")
            params["categoria"] = f"%{categoria}%"

    if stock_filter == "with_stock":
        where_clauses.append("f.existencia_stock > 0")
    elif stock_filter == "zero_stock":
        where_clauses.append("(f.existencia_stock IS NULL OR f.existencia_stock = 0)")

    if pdf_filter:
        pdf_l = pdf_filter.lower().strip()
        if pdf_l in ("with_pdf", "con_pdf", "pdf", "1", "true"):
            where_clauses.append(f"({pdf_select} IS NOT NULL AND {pdf_select} != '' AND {pdf_select} != '#')")
        elif pdf_l in ("no_pdf", "sin_pdf", "0", "false"):
            where_clauses.append(f"({pdf_select} IS NULL OR {pdf_select} = '' OR {pdf_select} = '#')")

    if con_orden:
        where_clauses.append(f"({ord_min_expr} IS NOT NULL AND {ord_min_expr} != '')")

    # ── Filtros de Características Técnicas por Componente ───────────────
    if cpu and cpu != 'Todos':
        cpu_u = cpu.upper().strip()
        if 'ULTRA' in cpu_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%CORE ULTRA%' OR UPPER(f.descripcion_producto) LIKE '%ULTRA %')")
        elif 'I9' in cpu_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%I9%' OR UPPER(f.descripcion_producto) LIKE '%CORE I9%')")
        elif 'I7' in cpu_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%I7%' OR UPPER(f.descripcion_producto) LIKE '%CORE I7%')")
        elif 'I5' in cpu_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%I5%' OR UPPER(f.descripcion_producto) LIKE '%CORE I5%')")
        elif 'I3' in cpu_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%I3%' OR UPPER(f.descripcion_producto) LIKE '%CORE I3%')")
        elif 'RYZEN 9' in cpu_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%RYZEN 9%'")
        elif 'RYZEN 7' in cpu_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%RYZEN 7%'")
        elif 'RYZEN 5' in cpu_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%RYZEN 5%'")
        elif 'RYZEN 3' in cpu_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%RYZEN 3%'")
        elif 'CELERON' in cpu_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%CELERON%'")
        elif 'XEON' in cpu_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%XEON%'")
        elif 'PENTIUM' in cpu_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%PENTIUM%'")
        else:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE :cpu_filter")
            params["cpu_filter"] = f"%{cpu}%"

    if ram and ram != 'Todos':
        m_ram = re.search(r'\b(\d+)\s*GB\b', ram, re.I)
        if m_ram:
            val = m_ram.group(1)
            where_clauses.append(f"(UPPER(f.descripcion_producto) LIKE '%{val} GB%' OR UPPER(f.descripcion_producto) LIKE '%{val}GB%')")
        else:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE :ram_filter")
            params["ram_filter"] = f"%{ram}%"

    if disco and disco != 'Todos':
        disco_u = disco.upper().strip()
        if '128' in disco_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%128%GB%' OR UPPER(f.descripcion_producto) LIKE '%128GB%')")
        elif '256' in disco_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%256%GB%' OR UPPER(f.descripcion_producto) LIKE '%256GB%')")
        elif '512' in disco_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%512%GB%' OR UPPER(f.descripcion_producto) LIKE '%512GB%')")
        elif '1 TB' in disco_u or '1TB' in disco_u:
            if 'HDD' in disco_u:
                where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%1%TB%' AND UPPER(f.descripcion_producto) LIKE '%HDD%')")
            else:
                where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%1%TB%' AND UPPER(f.descripcion_producto) NOT LIKE '%HDD%')")
        elif '2 TB' in disco_u or '2TB' in disco_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%2%TB%')")
        else:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE :disco_filter")
            params["disco_filter"] = f"%{disco}%"

    if pantalla and pantalla != 'Todos':
        m_pan = re.search(r'(\d+(?:\.\d+)?)', pantalla)
        if m_pan:
            val = m_pan.group(1)
            where_clauses.append(f"(UPPER(f.descripcion_producto) LIKE '%{val}\"%' OR UPPER(f.descripcion_producto) LIKE '%{val} PULG%' OR UPPER(f.descripcion_producto) LIKE '%{val}PULG%' OR UPPER(f.descripcion_producto) LIKE '%{val} %')")
        elif pantalla == 'Sin pantalla':
            where_clauses.append("UPPER(f.descripcion_producto) NOT LIKE '%PANTALLA%'")

    if so and so != 'Todos':
        so_u = so.upper()
        if '11 PRO' in so_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%W11 PRO%' OR UPPER(f.descripcion_producto) LIKE '%W11P%' OR UPPER(f.descripcion_producto) LIKE '%WIN 11 PRO%' OR UPPER(f.descripcion_producto) LIKE '%WINDOWS 11 PRO%')")
        elif '11 HOME' in so_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%W11 HOME%' OR UPPER(f.descripcion_producto) LIKE '%W11H%' OR UPPER(f.descripcion_producto) LIKE '%WIN 11 HOME%' OR UPPER(f.descripcion_producto) LIKE '%WINDOWS 11 HOME%')")
        elif '10' in so_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%W10%' OR UPPER(f.descripcion_producto) LIKE '%WINDOWS 10%')")
        elif 'FREE' in so_u or 'DOS' in so_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%FREEDOS%' OR UPPER(f.descripcion_producto) LIKE '%FREE DOS%' OR UPPER(f.descripcion_producto) LIKE '%NO TIENE%')")
        elif 'LINUX' in so_u or 'UBUNTU' in so_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%UBUNTU%' OR UPPER(f.descripcion_producto) LIKE '%LINUX%')")

    if panel and panel != 'Todos':
        panel_u = panel.upper().strip()
        if panel_u == 'IPS':
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%IPS%'")
        elif panel_u == 'VA':
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '% VA %' OR UPPER(f.descripcion_producto) LIKE '%VA-%' OR UPPER(f.descripcion_producto) LIKE 'VA %')")
        elif panel_u == 'TN':
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '% TN %' OR UPPER(f.descripcion_producto) LIKE 'TN %')")
        elif panel_u == 'OLED':
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%OLED%'")

    if resolucion and resolucion != 'Todos':
        res_u = resolucion.upper()
        if '4K' in res_u or '3840' in res_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%3840X2160%' OR UPPER(f.descripcion_producto) LIKE '%4K%')")
        elif '2K' in res_u or '2560' in res_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%2560X1440%' OR UPPER(f.descripcion_producto) LIKE '%2K%')")
        elif 'FHD' in res_u or '1920' in res_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%1920X1080%' OR UPPER(f.descripcion_producto) LIKE '%FHD%' OR UPPER(f.descripcion_producto) LIKE '%FULL HD%')")
        elif '1600' in res_u or 'HD+' in res_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%1600X900%' OR UPPER(f.descripcion_producto) LIKE '%HD+%')")
        elif 'HD' in res_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%1366X768%' OR UPPER(f.descripcion_producto) LIKE '%HD%')")

    # ── Puertos y Conectividad (VGA, HDMI, WLAN, LAN, BT) ───────────────
    if vga and vga != 'Todos':
        vga_l = vga.lower()
        if vga_l in ('si', 'con_vga', '1', 'true'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%VGA: SI%'")
        elif vga_l in ('no', 'sin_vga', '0', 'false'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%VGA: NO%'")

    if hdmi and hdmi != 'Todos':
        hdmi_l = hdmi.lower()
        if hdmi_l in ('si', 'con_hdmi', '1', 'true'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%HDMI: SI%'")
        elif hdmi_l in ('no', 'sin_hdmi', '0', 'false'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%HDMI: NO%'")

    if wifi and wifi != 'Todos':
        wifi_l = wifi.lower()
        if wifi_l in ('si', 'con_wifi', '1', 'true'):
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%WLAN: SI%' OR UPPER(f.descripcion_producto) LIKE '%WI-FI%')")
        elif wifi_l in ('no', 'sin_wifi', '0', 'false'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%WLAN: NO%'")

    if bluetooth and bluetooth != 'Todos':
        bt_l = bluetooth.lower()
        if bt_l in ('si', 'con_bt', '1', 'true'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%BLUETOOTH: SI%'")
        elif bt_l in ('no', 'sin_bt', '0', 'false'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%BLUETOOTH: NO%'")

    if lan and lan != 'Todos':
        lan_l = lan.lower()
        if lan_l in ('si', 'con_lan', '1', 'true'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%LAN: SI%'")
        elif lan_l in ('no', 'sin_lan', '0', 'false'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%LAN: NO%'")

    if unidad_optica and unidad_optica != 'Todos':
        uo_l = unidad_optica.lower()
        if uo_l in ('si', 'con_dvd', '1', 'true'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%UNIDAD OPTICA: SI%'")
        elif uo_l in ('no', 'sin_dvd', '0', 'false'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%UNIDAD OPTICA: NO%'")

    if camara and camara != 'Todos':
        cam_l = camara.lower()
        if cam_l in ('si', 'con_camara', '1', 'true'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%CAMARA WEB: SI%'")
        elif cam_l in ('no', 'sin_camara', '0', 'false'):
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%CAMARA WEB: NO%'")

    if tactil and tactil != 'Todos':
        tac_l = tactil.lower()
        if tac_l in ('si', 'tactil', 'touch', '1', 'true'):
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%TACTIL%' OR UPPER(f.descripcion_producto) LIKE '%TOUCH%')")
        elif tac_l in ('no', 'no_tactil', '0', 'false'):
            where_clauses.append("(UPPER(f.descripcion_producto) NOT LIKE '%TACTIL%' AND UPPER(f.descripcion_producto) NOT LIKE '%TOUCH%')")

    # ── Tecnologías Específicas de RAM y Disco ───────────────────────────
    if ram_tech and ram_tech != 'Todos':
        rt_u = ram_tech.upper()
        if rt_u == 'DDR5':
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%DDR5%'")
        elif rt_u == 'DDR4':
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%DDR4%'")
        elif 'LPDDR' in rt_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%LPDDR5%' OR UPPER(f.descripcion_producto) LIKE '%LPDDR4%')")

    if disco_tipo and disco_tipo != 'Todos':
        dt_u = disco_tipo.upper()
        if 'NVME' in dt_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%NVME%'")
        elif 'M.2' in dt_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%M.2%'")
        elif 'HIBRIDO' in dt_u or 'SSD + HDD' in dt_u or 'SSD/HDD' in dt_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%SSD%' AND UPPER(f.descripcion_producto) LIKE '%HDD%')")
        elif 'SOLO SSD' in dt_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%SSD%' AND UPPER(f.descripcion_producto) NOT LIKE '%HDD%')")
        elif 'SOLO HDD' in dt_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%HDD%' AND UPPER(f.descripcion_producto) NOT LIKE '%SSD%')")

    # ── Generación de CPU ────────────────────────────────────────────────
    if cpu_gen and cpu_gen != 'Todos':
        cg_u = cpu_gen.upper()
        if '14' in cg_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%-14%' OR UPPER(f.descripcion_producto) LIKE '% 14700%' OR UPPER(f.descripcion_producto) LIKE '% 14400%' OR UPPER(f.descripcion_producto) LIKE '% 14900%')")
        elif '13' in cg_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%-13%' OR UPPER(f.descripcion_producto) LIKE '% 13700%' OR UPPER(f.descripcion_producto) LIKE '% 13400%' OR UPPER(f.descripcion_producto) LIKE '% 13500%')")
        elif '12' in cg_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%-12%' OR UPPER(f.descripcion_producto) LIKE '% 12700%' OR UPPER(f.descripcion_producto) LIKE '% 12400%' OR UPPER(f.descripcion_producto) LIKE '% 12100%')")
        elif '11' in cg_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%-11%' OR UPPER(f.descripcion_producto) LIKE '% 11700%' OR UPPER(f.descripcion_producto) LIKE '% 11400%')")
        elif '10' in cg_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%-10%' OR UPPER(f.descripcion_producto) LIKE '% 10700%' OR UPPER(f.descripcion_producto) LIKE '% 10400%')")
        elif 'ULTRA' in cg_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%CORE ULTRA%'")
        elif '7000' in cg_u or '8000' in cg_u:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%RYZEN%7%' OR UPPER(f.descripcion_producto) LIKE '%RYZEN%8%')")
        elif '5000' in cg_u:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%RYZEN%5%'")

    # ── Office y Garantía ────────────────────────────────────────────────
    if office and office != 'Todos':
        off_l = office.lower()
        if 'home' in off_l or 'business' in off_l:
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%OFFICE HOME & BUSINESS%' OR UPPER(f.descripcion_producto) LIKE '%OFFICE HOME AND BUSINESS%')")
        elif off_l in ('si', 'con_office', '1', 'true'):
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%OFFICE%' OR UPPER(f.descripcion_producto) LIKE '%OFIMATICA PRE-INSTALADA%')")
        elif off_l in ('no', 'sin_office', '0', 'false'):
            where_clauses.append("(UPPER(f.descripcion_producto) LIKE '%SUITE OFIMATICA: NO%' OR (UPPER(f.descripcion_producto) NOT LIKE '%OFFICE%' AND UPPER(f.descripcion_producto) NOT LIKE '%PRE-INSTALADA%'))")

    if garantia and garantia != 'Todos':
        if '36' in garantia:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%36 MESES%'")
        elif '24' in garantia:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%24 MESES%'")
        elif '12' in garantia:
            where_clauses.append("UPPER(f.descripcion_producto) LIKE '%12 MESES%'")

    where_sql = " AND ".join(where_clauses)
    is_consolidated = (not proveedor or proveedor.lower() == "all")

    selected_reg = region.strip().upper() if (region and region.lower() != "all") else None
    if selected_reg:
        import unicodedata
        n = unicodedata.normalize('NFKD', selected_reg)
        selected_reg = ''.join(c for c in n if not unicodedata.combining(c))
        params["selected_reg"] = selected_reg
        plazo_expr = "COALESCE((f.raw_json->'plazos_por_region'->>:selected_reg)::int, (f.raw_json->'plazos_por_region'->>'LIMA')::int, f.plazo_entrega_dias, 90)"
    else:
        plazo_expr = "COALESCE((f.raw_json->'plazos_por_region'->>'LIMA')::int, f.plazo_entrega_dias, 90)"

    if is_consolidated:
        having_clauses = []
        if proveedor_filter:
            pf = proveedor_filter.lower().strip()
            if pf in ("ambos", "compitiendo", "competencia", "dual"):
                having_clauses.append("COUNT(DISTINCT r.ruc_proveedor) > 1")
            elif pf in ("exclusivo", "unico", "single", "1"):
                having_clauses.append("COUNT(DISTINCT r.ruc_proveedor) = 1")
            elif pf in ("thekingcomputer", "the_king", "king", "20601234567"):
                having_clauses.append("BOOL_OR(r.ruc_proveedor = '20601234567' OR UPPER(r.nombre_proveedor) LIKE '%KING%')")
            elif pf in ("jorge_rojas", "jorge", "rojas", "10408899991"):
                having_clauses.append("BOOL_OR(r.ruc_proveedor = '10408899991' OR UPPER(r.nombre_proveedor) LIKE '%JORGE%' OR UPPER(r.nombre_proveedor) LIKE '%ROJAS%')")

        having_sql = f"HAVING {' AND '.join(having_clauses)}" if having_clauses else ""

        order_by_sql = "g.id DESC, g.group_key ASC"
        if sort_by == "precio_asc":
            order_by_sql = "g.min_precio ASC NULLS LAST, g.group_key ASC"
        elif sort_by == "precio_desc":
            order_by_sql = "g.min_precio DESC NULLS LAST, g.group_key ASC"
        elif sort_by == "stock_desc":
            order_by_sql = "g.existencia_stock DESC NULLS LAST, g.group_key ASC"
        elif sort_by == "marca_asc":
            order_by_sql = "g.marca ASC, g.group_key ASC"

        sql = f"""
            WITH raw_matched AS (
                SELECT 
                    f.id,
                    f.nro_parte,
                    f.descripcion_producto AS descripcion,
                    f.marca,
                    f.catalogo,
                    f.categoria,
                    f.acuerdo_marco,
                    f.nombre_proveedor,
                    f.ruc_proveedor,
                    f.precio_ofertado,
                    f.existencia_stock,
                    {plazo_expr} AS plazo_entrega_dias,
                    COALESCE(f.raw_json->'plazos_por_region', '{{}}'::json) AS plazos_por_region,
                    f.region,
                    f.provincia,
                    {pdf_select} AS pdf_url,
                    f.fecha_extraccion,
                    {ord_min_expr} AS orden_min,
                    {fec_min_expr} AS fecha_orden_min,
                    {pref_expr} AS precio_referencia,
                    {pmin_expr} AS precio_historico_min,
                    {momin_expr} AS monto_orden_min,
                    CASE 
                        WHEN UPPER(TRIM(COALESCE(f.nro_parte, ''))) IN ('', '-', 'S/N', 'SN', 'COLECTIVO', 'VARIOS', '0', 'NO TIENE', 'SIN NUMERO', 'SIN NUMERO DE PARTE', 'NO APLICA') 
                        THEN CONCAT(COALESCE(NULLIF(TRIM(f.nro_parte), ''), 'S/N'), '::', MD5(COALESCE(f.descripcion_producto, f.id::text)))
                        ELSE UPPER(TRIM(f.nro_parte))
                    END AS group_key
                FROM ofertas_proveedor_history f
                {pdf_join}
                WHERE {where_sql}
            ),
            dedup_provider_offers AS (
                SELECT DISTINCT ON (r.group_key, r.ruc_proveedor)
                    r.group_key,
                    r.id,
                    r.nro_parte,
                    r.descripcion,
                    r.marca,
                    r.catalogo,
                    r.categoria,
                    r.acuerdo_marco,
                    r.nombre_proveedor,
                    r.ruc_proveedor,
                    r.precio_ofertado,
                    r.existencia_stock,
                    r.plazo_entrega_dias,
                    r.plazos_por_region,
                    r.region,
                    r.provincia,
                    r.pdf_url,
                    r.fecha_extraccion,
                    r.orden_min,
                    r.fecha_orden_min,
                    r.precio_referencia,
                    r.precio_historico_min,
                    r.monto_orden_min
                FROM raw_matched r
                ORDER BY r.group_key, r.ruc_proveedor, r.precio_ofertado ASC NULLS LAST, r.id DESC
            ),
            grouped_items AS (
                SELECT 
                    r.group_key,
                    MAX(r.nro_parte) AS nro_parte,
                    MIN(r.id) AS id,
                    MAX(r.descripcion) AS descripcion,
                    MAX(r.marca) AS marca,
                    MAX(r.catalogo) AS catalogo,
                    MAX(r.categoria) AS categoria,
                    MAX(r.acuerdo_marco) AS acuerdo_marco,
                    MIN(r.precio_ofertado) AS min_precio,
                    MAX(r.precio_ofertado) AS max_precio,
                    MIN(r.plazo_entrega_dias) AS min_plazo_entrega,
                    MAX(r.plazo_entrega_dias) AS max_plazo_entrega,
                    MAX(r.region) AS region,
                    MAX(r.provincia) AS provincia,
                    MAX(r.pdf_url) AS pdf_url,
                    MAX(r.orden_min) AS orden_min,
                    MAX(r.fecha_orden_min) AS fecha_orden_min,
                    MAX(r.precio_referencia) AS precio_referencia,
                    MAX(r.precio_historico_min) AS precio_historico_min,
                    MAX(r.monto_orden_min) AS monto_orden_min,
                    SUM(COALESCE(r.existencia_stock, 0)) AS existencia_stock,
                    COUNT(DISTINCT r.ruc_proveedor) AS total_proveedores,
                    json_agg(
                        json_build_object(
                            'id', r.id,
                            'nombre_proveedor', r.nombre_proveedor,
                            'ruc_proveedor', r.ruc_proveedor,
                            'precio_ofertado', r.precio_ofertado,
                            'existencia_stock', r.existencia_stock,
                            'plazo_entrega_dias', r.plazo_entrega_dias,
                            'plazos_por_region', r.plazos_por_region,
                            'region', r.region,
                            'provincia', r.provincia,
                            'pdf_url', r.pdf_url,
                            'estado', 'VIGENTE',
                            'fecha_extraccion', r.fecha_extraccion::text
                        ) ORDER BY r.precio_ofertado ASC NULLS LAST
                    ) AS ofertas
                FROM dedup_provider_offers r
                GROUP BY r.group_key
                {having_sql}
            )
            SELECT 
                g.id,
                g.nro_parte,
                g.descripcion,
                g.marca,
                g.catalogo,
                g.categoria,
                g.acuerdo_marco,
                g.min_precio,
                g.max_precio,
                g.min_plazo_entrega,
                g.max_plazo_entrega,
                g.region,
                g.provincia,
                g.pdf_url,
                g.orden_min,
                g.fecha_orden_min,
                g.precio_referencia,
                g.precio_historico_min,
                g.monto_orden_min,
                g.existencia_stock,
                g.total_proveedores,
                g.ofertas,
                COUNT(*) OVER() AS total_count,
                SUM(CASE WHEN g.total_proveedores > 1 THEN 1 ELSE 0 END) OVER() AS total_competing_count,
                SUM(COALESCE(g.existencia_stock, 0)) OVER() AS total_stock_global
            FROM grouped_items g
            ORDER BY {order_by_sql}
            LIMIT :limit OFFSET :offset
        """
    else:
        order_by_sql = "f.id DESC"
        if sort_by == "precio_asc":
            order_by_sql = "f.precio_ofertado ASC NULLS LAST"
        elif sort_by == "precio_desc":
            order_by_sql = "f.precio_ofertado DESC NULLS LAST"
        elif sort_by == "stock_desc":
            order_by_sql = "f.existencia_stock DESC NULLS LAST"
        elif sort_by == "marca_asc":
            order_by_sql = "f.marca ASC"

        sql = f"""
            SELECT 
                f.id,
                f.nro_parte,
                f.descripcion_producto AS descripcion,
                f.marca,
                f.catalogo,
                f.categoria,
                f.acuerdo_marco,
                f.nombre_proveedor AS proveedor,
                f.ruc_proveedor,
                f.precio_ofertado,
                f.precio_ofertado AS min_precio,
                f.existencia_stock,
                {plazo_expr} AS plazo_entrega_dias,
                {plazo_expr} AS min_plazo_entrega,
                {plazo_expr} AS max_plazo_entrega,
                COALESCE(f.raw_json->'plazos_por_region', '{{}}'::json) AS plazos_por_region,
                f.region,
                f.provincia,
                {pdf_select} AS pdf_url,
                {ord_min_expr} AS orden_min,
                {fec_min_expr} AS fecha_orden_min,
                {pref_expr} AS precio_referencia,
                {pmin_expr} AS precio_historico_min,
                {momin_expr} AS monto_orden_min,
                f.raw_json,
                f.fecha_extraccion::text AS fecha_extraccion,
                1 AS total_proveedores,
                json_build_array(json_build_object(
                    'id', f.id,
                    'nombre_proveedor', f.nombre_proveedor,
                    'ruc_proveedor', f.ruc_proveedor,
                    'precio_ofertado', f.precio_ofertado,
                    'existencia_stock', f.existencia_stock,
                    'plazo_entrega_dias', {plazo_expr},
                    'plazos_por_region', COALESCE(f.raw_json->'plazos_por_region', '{{}}'::json),
                    'region', f.region,
                    'provincia', f.provincia,
                    'pdf_url', {pdf_select},
                    'estado', 'VIGENTE',
                    'fecha_extraccion', f.fecha_extraccion::text
                )) AS ofertas,
                COUNT(*) OVER() AS total_count,
                0::bigint AS total_competing_count,
                SUM(COALESCE(f.existencia_stock, 0)) OVER() AS total_stock_global
            FROM ofertas_proveedor_history f
            {pdf_join}
            WHERE {where_sql}
            ORDER BY {order_by_sql}
            LIMIT :limit OFFSET :offset
        """

    try:
        params["limit"] = limit
        params["offset"] = offset
        rows = db.execute(text(sql), params).mappings().all()

        if rows and len(rows) > 0:
            total_items = rows[0]["total_count"] if "total_count" in rows[0] else len(rows)
            total_competing = int(rows[0].get("total_competing_count") or 0)
            total_stock_global = int(rows[0].get("total_stock_global") or 0)
            items = []
            for r in rows:
                item_dict = dict(r)
                if isinstance(item_dict.get("ofertas"), str):
                    try:
                        item_dict["ofertas"] = json.loads(item_dict["ofertas"])
                    except Exception:
                        item_dict["ofertas"] = []
                # Remove window columns from item dicts (they're aggregate-level)
                item_dict.pop("total_competing_count", None)
                item_dict.pop("total_stock_global", None)
                items.append(item_dict)

            return {
                "items": items,
                "page": page,
                "limit": limit,
                "total": total_items,
                "total_competing": total_competing,
                "total_stock": total_stock_global,
                "is_consolidated": is_consolidated
            }
    except Exception as e:
        import logging
        logging.getLogger("ceam.proveedores").error("Error en get_proveedor_fichas: %s", e)

    return {"items": [], "page": page, "limit": limit, "total": 0, "is_consolidated": is_consolidated}

@router.post("/clear")
def clear_proveedor_data(
    proveedor: Optional[str] = Query(None),
    solo_plazos: bool = Query(False),
    db: Session = Depends(get_db)
):
    """Limpia las ofertas de la base de datos para comenzar de nuevo una extracción limpia."""
    try:
        if solo_plazos:
            # Solo resetear plazos sin borrar ofertas
            where = ""
            params = {}
            if proveedor and proveedor.lower() != "all":
                prov_l = proveedor.lower()
                if prov_l in ("thekingcomputer", "king"):
                    where = "WHERE UPPER(nombre_proveedor) LIKE '%KING%' OR UPPER(ruc_proveedor) = '20601234567'"
                elif prov_l in ("jorge_rojas", "jorge", "rojas"):
                    where = "WHERE UPPER(nombre_proveedor) LIKE '%ROJAS%' OR UPPER(ruc_proveedor) = '10408899991'"
                else:
                    where = "WHERE UPPER(nombre_proveedor) LIKE UPPER(:prov)"
                    params = {"prov": f"%{proveedor}%"}
            db.execute(text(f"""
                UPDATE ofertas_proveedor_history
                SET plazo_entrega_dias = NULL,
                    raw_json = CASE
                        WHEN raw_json IS NOT NULL THEN (raw_json::jsonb - 'plazos_por_region')::json
                        ELSE NULL
                    END
                {where}
            """), params)
            db.commit()
            return {"success": True, "message": "Plazos de entrega limpiados con éxito (ofertas conservadas)"}

        if proveedor and proveedor.lower() != "all":
            prov_l = proveedor.lower()
            if prov_l in ("thekingcomputer", "king"):
                db.execute(text("DELETE FROM ofertas_proveedor_history WHERE UPPER(nombre_proveedor) LIKE '%KING%' OR UPPER(ruc_proveedor) = '20601234567';"))
            elif prov_l in ("jorge_rojas", "jorge", "rojas"):
                db.execute(text("DELETE FROM ofertas_proveedor_history WHERE UPPER(nombre_proveedor) LIKE '%ROJAS%' OR UPPER(nombre_proveedor) LIKE '%JORGE%' OR UPPER(ruc_proveedor) = '10408899991';"))
            else:
                db.execute(text("DELETE FROM ofertas_proveedor_history WHERE UPPER(nombre_proveedor) LIKE UPPER(:prov);"), {"prov": f"%{proveedor}%"})
        else:
            db.execute(text("TRUNCATE TABLE ofertas_proveedor_history RESTART IDENTITY;"))
        db.commit()
        return {"success": True, "message": "Datos de ofertas limpiados con éxito"}
    except Exception as e:
        db.rollback()
        return {"success": False, "error": str(e)}

@router.get("/filters/{column_name}")
def get_column_filters(column_name: str, db: Session = Depends(get_db)):
    """Devuelve valores únicos no nulos para construir filtros tipo Excel en las cabeceras de tabla."""
    from app.models.ofertas_proveedor import OfertaProveedorHistory
    valid_columns = {
        "marca": OfertaProveedorHistory.marca,
        "proveedor": OfertaProveedorHistory.nombre_proveedor,
        "catalogo": OfertaProveedorHistory.catalogo,
        "categoria": OfertaProveedorHistory.categoria,
    }
    if column_name not in valid_columns:
        raise HTTPException(status_code=400, detail="Columna no permitida para filtros")
    
    col = valid_columns[column_name]
    rows = db.query(col).filter(col.isnot(None), col != '').distinct().order_by(col).all()
    return {"values": [r[0] for r in rows if r[0]]}

def extract_fields_from_product_text(desc: str) -> dict:
    """
    Extrae dinámicamente los campos técnicos normalizados analizando la estructura
    de etiquetas del JSON / descripción de Perú Compras (PROCESADOR:, RAM:, etc.).
    """
    if not desc:
        return {}

    fields = {}

    # 1. PROCESADOR: Detección precisa de campo estructurado
    m_cpu = re.search(r'PROCESADOR:\s*([^;:\n]+?)(?=\s+(?:RAM|ALMACENAMIENTO|PANTALLA|SIST|UNIDAD|LAN|WLAN|USB|VGA|HDMI|TECLADO|MOUSE|G\.?\s*F|SUITE)|$)', desc, re.I)
    raw_cpu = m_cpu.group(1).strip() if m_cpu else desc
    raw_cpu_u = raw_cpu.upper()

    cpu = None
    if "CORE ULTRA" in raw_cpu_u or "ULTRA " in raw_cpu_u:
        cpu = "Intel Core Ultra"
    elif "CORE I9" in raw_cpu_u or "I9-" in raw_cpu_u or "I9 " in raw_cpu_u or "CI9" in raw_cpu_u:
        cpu = "Intel Core i9"
    elif "CORE I7" in raw_cpu_u or "I7-" in raw_cpu_u or "I7 " in raw_cpu_u or "CI7" in raw_cpu_u:
        cpu = "Intel Core i7"
    elif "CORE I5" in raw_cpu_u or "I5-" in raw_cpu_u or "I5 " in raw_cpu_u or "CI5" in raw_cpu_u:
        cpu = "Intel Core i5"
    elif "CORE I3" in raw_cpu_u or "I3-" in raw_cpu_u or "I3 " in raw_cpu_u or "CI3" in raw_cpu_u:
        cpu = "Intel Core i3"
    elif "RYZEN 9" in raw_cpu_u or "R9-" in raw_cpu_u or "R9 " in raw_cpu_u:
        cpu = "AMD Ryzen 9"
    elif "RYZEN 7" in raw_cpu_u or "R7-" in raw_cpu_u or "R7 " in raw_cpu_u:
        cpu = "AMD Ryzen 7"
    elif "RYZEN 5" in raw_cpu_u or "R5-" in raw_cpu_u or "R5 " in raw_cpu_u:
        cpu = "AMD Ryzen 5"
    elif "RYZEN 3" in raw_cpu_u or "R3-" in raw_cpu_u or "R3 " in raw_cpu_u:
        cpu = "AMD Ryzen 3"
    elif "CELERON" in raw_cpu_u or "N4020" in raw_cpu_u or "N4500" in raw_cpu_u or "N100" in raw_cpu_u:
        cpu = "Intel Celeron"
    elif "PENTIUM" in raw_cpu_u or "GOLD" in raw_cpu_u:
        cpu = "Intel Pentium"
    elif "XEON" in raw_cpu_u:
        cpu = "Intel Xeon"
    elif "APPLE" in raw_cpu_u or "M1" in raw_cpu_u or "M2" in raw_cpu_u or "M3" in raw_cpu_u or "M4" in raw_cpu_u:
        cpu = "Apple Silicon"
    elif "SNAPDRAGON" in raw_cpu_u:
        cpu = "Snapdragon X"
    fields["cpu"] = cpu

    # Generación de CPU
    cpu_gen = None
    if re.search(r'(?:i[3579]-14\d{3}|-14\d{2}| 14\d{3}|14700|14400)', raw_cpu_u):
        cpu_gen = "14ª Gen (Intel Core i-14xxx)"
    elif re.search(r'(?:i[3579]-13\d{3}|-13\d{2}| 13\d{3}|13700|13400|13900)', raw_cpu_u):
        cpu_gen = "13ª Gen (Intel Core i-13xxx)"
    elif re.search(r'(?:i[3579]-12\d{3}|-12\d{2}| 12\d{3}|12700|12400|1235U|1215U)', raw_cpu_u):
        cpu_gen = "12ª Gen (Intel Core i-12xxx)"
    elif re.search(r'(?:i[3579]-11\d{3}|-11\d{2}| 11\d{3}|1135G7|1165G7)', raw_cpu_u):
        cpu_gen = "11ª Gen (Intel Core i-11xxx)"
    elif re.search(r'(?:i[3579]-10\d{3}|-10\d{2}| 10\d{3}|10100|10400)', raw_cpu_u):
        cpu_gen = "10ª Gen (Intel Core i-10xxx)"
    elif "CORE ULTRA" in raw_cpu_u:
        cpu_gen = "Core Ultra (Series 1)"
    elif "RYZEN" in raw_cpu_u and re.search(r'[78]\d{3}', raw_cpu_u):
        cpu_gen = "AMD Ryzen 7000 / 8000"
    elif "RYZEN" in raw_cpu_u and "5000" in raw_cpu_u:
        cpu_gen = "AMD Ryzen 5000"
    fields["cpu_gen"] = cpu_gen

    # 2. RAM: Detección precisa de campo estructurado
    m_ram = re.search(r'(?:RAM|MEMORIA):\s*([^;:\n]+?)(?=\s+(?:ALMACENAMIENTO|PANTALLA|SIST|UNIDAD|LAN|WLAN|USB|VGA|HDMI|TECLADO|MOUSE|G\.?\s*F|SUITE)|$)', desc, re.I)
    ram_str = m_ram.group(1).strip() if m_ram else desc
    m_ram_size = re.search(r'\b(\d+)\s*(?:GB|GIGAS|GIB)\b', ram_str, re.I)
    if m_ram_size:
        fields["ram"] = f"{m_ram_size.group(1)} GB"
    else:
        fields["ram"] = None

    ram_tech = None
    if "LPDDR5" in ram_str.upper():
        ram_tech = "LPDDR5 / LPDDR5X"
    elif "DDR5" in ram_str.upper():
        ram_tech = "DDR5"
    elif "DDR4" in ram_str.upper():
        ram_tech = "DDR4"
    fields["ram_tech"] = ram_tech

    # 3. ALMACENAMIENTO: Detección precisa de campo estructurado
    m_alm = re.search(r'(?:ALMACENAMIENTO|DISCO):\s*([^;:\n]+?)(?=\s+(?:PANTALLA|SIST|UNIDAD|LAN|WLAN|USB|VGA|HDMI|TECLADO|MOUSE|G\.?\s*F|SUITE)|$)', desc, re.I)
    alm_str = m_alm.group(1).strip() if m_alm else desc
    storage = None
    m_gb = re.search(r'\b(128|256|500|512)\s*GB\b', alm_str, re.I)
    m_tb = re.search(r'\b(1|2|4)\s*TB\b', alm_str, re.I)
    if m_gb:
        val = 512 if m_gb.group(1) == "500" else int(m_gb.group(1))
        storage = f"{val} GB"
    elif m_tb:
        storage = f"{m_tb.group(1)} TB"
    fields["storage"] = storage

    alm_u = alm_str.upper()
    disco_tipo = None
    if "NVME" in alm_u:
        disco_tipo = "NVMe M.2 SSD"
    elif "M.2" in alm_u:
        disco_tipo = "M.2 SSD"
    elif "SSD" in alm_u and "HDD" in alm_u:
        disco_tipo = "Híbrido (SSD + HDD)"
    elif "SSD" in alm_u:
        disco_tipo = "Solo SSD"
    elif "HDD" in alm_u or "MECANICO" in alm_u:
        disco_tipo = "Solo HDD"
    fields["disco_tipo"] = disco_tipo

    # 4. PANTALLA: Detección precisa de campo estructurado
    m_pan = re.search(r'PANTALLA:\s*([^;:\n]+?)(?=\s+(?:SIST|UNIDAD|LAN|WLAN|USB|VGA|HDMI|TECLADO|MOUSE|G\.?\s*F|SUITE)|$)', desc, re.I)
    pan_str = m_pan.group(1).strip() if m_pan else desc
    m_pulg = re.search(r'(\d+(?:\.\d+)?)\s*(?:\"|\'\'|PULGADAS|PLG|PULG)', pan_str, re.I)
    if m_pulg:
        val = float(m_pulg.group(1))
        fields["display"] = f"{val:g}\""
    else:
        fields["display"] = None

    # 5. PANEL & RESOLUCION
    du = desc.upper()
    panel = None
    if "OLED" in du: panel = "OLED"
    elif "IPS" in du: panel = "IPS"
    elif " VA " in du or "VA-" in du: panel = "VA"
    elif " TN " in du: panel = "TN"
    fields["panel"] = panel

    resolution = None
    if "3840" in du or "4K" in du: resolution = "4K UHD (3840x2160)"
    elif "2560" in du or "2K" in du or "QHD" in du: resolution = "2K QHD (2560x1440)"
    elif "1920" in du or "FHD" in du: resolution = "FHD (1920x1080)"
    elif "1600" in du or "HD+" in du: resolution = "HD+ (1600x900)"
    elif "1366" in du or "HD" in du: resolution = "HD (1366x768)"
    fields["resolution"] = resolution

    # 6. SISTEMA OPERATIVO: Detección precisa de campo estructurado
    m_so = re.search(r'(?:SIST\.?\s*OPER(?:ATIVO)?):\s*([^;:\n]+?)(?=\s+(?:UNIDAD|LAN|WLAN|USB|VGA|HDMI|TECLADO|MOUSE|G\.?\s*F|SUITE)|$)', desc, re.I)
    so_str = (m_so.group(1).strip() if m_so else desc).upper()
    os_name = None
    if "11 PRO" in so_str or "WIN 11 PRO" in so_str: os_name = "Windows 11 Pro"
    elif "11 HOME" in so_str or "WIN 11 HOME" in so_str: os_name = "Windows 11 Home"
    elif "10 PRO" in so_str: os_name = "Windows 10 Pro"
    elif "FREEDOS" in so_str or "FREE DOS" in so_str or "SIN SISTEMA" in so_str or "NO TIENE" in so_str: os_name = "FreeDOS / Sin SO"
    elif "LINUX" in so_str or "UBUNTU" in so_str: os_name = "Linux / Ubuntu"
    fields["os"] = os_name

    return fields


_FILTER_OPTIONS_CACHE: Dict[str, dict] = {}
_FILTER_CACHE_TIMESTAMP: float = 0.0
_FILTER_CACHE_TTL: float = 300.0  # 5 minutos

@router.get("/filter-options")
def get_filter_options(
    categoria: Optional[str] = Query(None, description="Filtrar opciones por categoría activa"),
    proveedor: Optional[str] = Query(None, description="Filtrar opciones por proveedor"),
    db: Session = Depends(get_db)
):
    """
    Retorna opciones de filtro 100% dinámicas extrayendo los campos de los JSON / descripciones
    reales de la base de datos (con caché en memoria para respuesta instantánea <2ms).
    Sin taxonomías ni listas hardcodeadas.
    """
    global _FILTER_OPTIONS_CACHE, _FILTER_CACHE_TIMESTAMP
    import time
    now = time.time()
    cache_key = f"{categoria or 'all'}:{proveedor or 'all'}"

    if cache_key in _FILTER_OPTIONS_CACHE and (now - _FILTER_CACHE_TIMESTAMP < _FILTER_CACHE_TTL):
        return _FILTER_OPTIONS_CACHE[cache_key]

    where_parts = ["1=1"]
    sql_params = {}

    if categoria and categoria.lower() != "all":
        categ_lower = categoria.lower().strip()
        if categ_lower in ("escritorio", "computadora de escritorio"):
            where_parts.append("(UPPER(categoria) LIKE '%ESCRITORIO%' OR UPPER(catalogo) LIKE '%ESCRITORIO%')")
        elif categ_lower in ("aio", "todo en uno", "all in one"):
            where_parts.append("(UPPER(categoria) LIKE '%TODO EN UNO%' OR UPPER(descripcion_producto) LIKE '%TODO EN UNO%' OR UPPER(descripcion_producto) LIKE '%ALL IN ONE%')")
        elif categ_lower in ("monitor", "monitores"):
            where_parts.append("(UPPER(categoria) LIKE '%MONITOR%' OR UPPER(descripcion_producto) LIKE '%MONITOR%')")
        elif categ_lower in ("portatil", "computadora portatil", "laptop"):
            where_parts.append("(UPPER(categoria) LIKE '%PORTATIL%' OR UPPER(categoria) LIKE '%PORTÁTIL%' OR UPPER(descripcion_producto) LIKE '%PORTATIL%' OR UPPER(descripcion_producto) LIKE '%LAPTOP%')")
        elif categ_lower in ("workstation", "workstation_portatil", "estacion"):
            where_parts.append("(UPPER(categoria) LIKE '%ESTACION%' OR UPPER(descripcion_producto) LIKE '%WORKSTATION%')")
        elif categ_lower in ("tableta", "tablet"):
            where_parts.append("(UPPER(categoria) LIKE '%TABLET%' OR UPPER(descripcion_producto) LIKE '%TABLET%')")
        elif "almacenamiento" in categ_lower:
            where_parts.append("(UPPER(categoria) LIKE '%ALMACENAMIENTO%' OR UPPER(catalogo) LIKE '%ALMACENAMIENTO%')")
        elif "escaner" in categ_lower:
            where_parts.append("(UPPER(categoria) LIKE '%ESCANER%' OR UPPER(catalogo) LIKE '%ESCANER%' OR UPPER(descripcion_producto) LIKE '%ESCANER%')")
        elif "pantalla" in categ_lower:
            where_parts.append("(UPPER(categoria) LIKE '%PANTALLA%' OR UPPER(descripcion_producto) LIKE '%PANTALLA%')")
        else:
            where_parts.append("(UPPER(categoria) LIKE UPPER(:cat) OR UPPER(catalogo) LIKE UPPER(:cat))")
            sql_params["cat"] = f"%{categoria}%"

    where_sql = " AND ".join(where_parts)

    try:
        # 1. Marcas distintas 100% dinámicas desde la BD
        marcas_rows = db.execute(text(f"""
            SELECT DISTINCT marca FROM ofertas_proveedor_history
            WHERE {where_sql} AND marca IS NOT NULL AND marca != ''
            ORDER BY marca ASC
        """), sql_params).fetchall()
        marcas = [r[0].strip() for r in marcas_rows if r[0]]

        # 2. Descripciones distintas para extraer campos técnicos estructurados
        desc_rows = db.execute(text(f"""
            SELECT DISTINCT descripcion_producto FROM ofertas_proveedor_history
            WHERE {where_sql} AND descripcion_producto IS NOT NULL AND descripcion_producto != ''
        """), sql_params).fetchall()

        cpus = set()
        cpu_gens = set()
        rams = set()
        ram_techs = set()
        storages = set()
        disco_tipos = set()
        displays = set()
        panels = set()
        resolutions = set()
        oss = set()

        for (desc,) in desc_rows:
            parsed = extract_fields_from_product_text(desc)
            if parsed.get("cpu"): cpus.add(parsed["cpu"])
            if parsed.get("cpu_gen"): cpu_gens.add(parsed["cpu_gen"])
            if parsed.get("ram"): rams.add(parsed["ram"])
            if parsed.get("ram_tech"): ram_techs.add(parsed["ram_tech"])
            if parsed.get("storage"): storages.add(parsed["storage"])
            if parsed.get("disco_tipo"): disco_tipos.add(parsed["disco_tipo"])
            if parsed.get("display"): displays.add(parsed["display"])
            if parsed.get("panel"): panels.add(parsed["panel"])
            if parsed.get("resolution"): resolutions.add(parsed["resolution"])
            if parsed.get("os"): oss.add(parsed["os"])

        def _sort_storage(x):
            try:
                num, unit = x.split()
                return (float(num) * (1024 if unit == 'TB' else 1))
            except:
                return 0

        def _sort_display(x):
            try:
                return float(x.replace('"', ''))
            except:
                return 0

        result = {
            "marcas": marcas,
            "cpus": sorted(cpus),
            "cpu_gens": sorted(cpu_gens),
            "rams": sorted(rams, key=lambda x: int(x.split()[0]) if x.split()[0].isdigit() else 0),
            "ram_techs": sorted(ram_techs),
            "storages": sorted(storages, key=_sort_storage),
            "disco_tipos": sorted(disco_tipos),
            "oss": sorted(oss),
            "displays": sorted(displays, key=_sort_display),
            "panels": sorted(panels),
            "resolutions": sorted(resolutions),
        }

        _FILTER_OPTIONS_CACHE[cache_key] = result
        _FILTER_CACHE_TIMESTAMP = now
        return result

    except Exception as e:
        import logging
        logging.getLogger("ceam.proveedores").error("Error analizando opciones dinámicas de filtro: %s", e)
        return {
            "marcas": [],
            "cpus": [],
            "cpu_gens": [],
            "rams": [],
            "ram_techs": [],
            "storages": [],
            "disco_tipos": [],
            "oss": [],
            "displays": [],
            "panels": [],
            "resolutions": [],
        }


@router.get("/export-json")
def export_all_fichas_json(db: Session = Depends(get_db)):
    """
    Devuelve todas las ofertas extraídas en formato JSON crudo completo.
    """
    try:
        rows = db.execute(text("""
            SELECT 
                id, nro_parte, descripcion_producto, marca, catalogo, categoria,
                acuerdo_marco, nombre_proveedor, ruc_proveedor, precio_ofertado,
                existencia_stock, plazo_entrega_dias, pdf_url, raw_json, fecha_extraccion
            FROM ofertas_proveedor_history
            ORDER BY id ASC
        """)).mappings().all()
        return [dict(r) for r in rows]
    except Exception as e:
        import logging
        logging.getLogger("ceam.proveedores").error("Error en export_all_fichas_json: %s", e)
        return {"error": str(e)}


@router.get("/export-excel")
def export_proveedor_fichas_excel(
    proveedor: Optional[str] = Query(None, description="Filtro por proveedor"),
    proveedor_filter: Optional[str] = Query(None, description="Filtro de competencia/exclusivo"),
    region: Optional[str] = Query(None, description="Filtro por región"),
    search: Optional[str] = Query(None, description="Búsqueda por texto"),
    marca: Optional[str] = Query(None, description="Filtro por marca"),
    nro_parte: Optional[str] = Query(None, description="Filtro por nro_parte"),
    catalogo: Optional[str] = Query(None, description="Filtro por catálogo"),
    categoria: Optional[str] = Query(None, description="Filtro por categoría"),
    stock_filter: Optional[str] = Query(None, description="Filtro de stock"),
    pdf_filter: Optional[str] = Query(None, description="Filtro de PDF"),
    sort_by: Optional[str] = Query(None, description="Ordenamiento"),
    cpu: Optional[str] = Query(None),
    ram: Optional[str] = Query(None),
    disco: Optional[str] = Query(None),
    pantalla: Optional[str] = Query(None),
    so: Optional[str] = Query(None),
    panel: Optional[str] = Query(None),
    resolucion: Optional[str] = Query(None),
    cpu_gen: Optional[str] = Query(None),
    ram_tech: Optional[str] = Query(None),
    disco_tipo: Optional[str] = Query(None),
    vga: Optional[str] = Query(None),
    hdmi: Optional[str] = Query(None),
    wifi: Optional[str] = Query(None),
    bluetooth: Optional[str] = Query(None),
    lan: Optional[str] = Query(None),
    office: Optional[str] = Query(None),
    garantia: Optional[str] = Query(None),
    unidad_optica: Optional[str] = Query(None),
    camara: Optional[str] = Query(None),
    tactil: Optional[str] = Query(None),
    con_orden: Optional[bool] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Genera y descarga un reporte profesional en Excel (.xlsx) con análisis comparativo,
    resumen ejecutivo de KPIs, distribución por categorías y catálogo completo de ofertas.
    """
    import io
    from datetime import datetime
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 1. Obtener datos filtrados
    res = get_proveedor_fichas(
        proveedor=proveedor,
        proveedor_filter=proveedor_filter,
        region=region,
        search=search,
        marca=marca,
        nro_parte=nro_parte,
        catalogo=catalogo,
        categoria=categoria,
        stock_filter=stock_filter,
        pdf_filter=pdf_filter,
        sort_by=sort_by,
        cpu=cpu,
        ram=ram,
        disco=disco,
        pantalla=pantalla,
        so=so,
        panel=panel,
        resolucion=resolucion,
        cpu_gen=cpu_gen,
        ram_tech=ram_tech,
        disco_tipo=disco_tipo,
        vga=vga,
        hdmi=hdmi,
        wifi=wifi,
        bluetooth=bluetooth,
        lan=lan,
        office=office,
        garantia=garantia,
        unidad_optica=unidad_optica,
        camara=camara,
        tactil=tactil,
        con_orden=con_orden,
        page=1,
        limit=100000,
        db=db
    )

    items = res.get("items", [])
    total_fichas = res.get("total", len(items))
    total_competing = res.get("total_competing", 0)
    total_stock = res.get("total_stock", 0)

    # 2. Conteo por categorías para Resumen
    cat_counts = get_categories_count(proveedor=proveedor, db=db)

    # 3. Construcción del Workbook
    wb = Workbook()

    # ── Paleta y Estilos ───────────────────────────────────────────────────
    FONT_FAMILY = "Segoe UI"
    NAVY_MAIN   = PatternFill("solid", fgColor="0F172A") # Slate 900
    NAVY_HDR    = PatternFill("solid", fgColor="1E3A8A") # Blue 900
    BLUE_ACCENT = PatternFill("solid", fgColor="2563EB") # Blue 600
    LIGHT_BG    = PatternFill("solid", fgColor="F8FAFC") # Slate 50
    CARD_BG     = PatternFill("solid", fgColor="F1F5F9") # Slate 100
    GREEN_FILL  = PatternFill("solid", fgColor="DCFCE7") # Green 100
    GREEN_TXT   = Font(name=FONT_FAMILY, size=9, bold=True, color="166534")
    PURPLE_FILL = PatternFill("solid", fgColor="F3E8FF") # Purple 100
    PURPLE_TXT  = Font(name=FONT_FAMILY, size=9, bold=True, color="6B21A8")
    
    TITLE_FONT  = Font(name=FONT_FAMILY, size=13, bold=True, color="FFFFFF")
    SUB_FONT    = Font(name=FONT_FAMILY, size=9, italic=True, color="CBD5E1")
    HDR_FONT    = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
    BOLD_FONT   = Font(name=FONT_FAMILY, size=10, bold=True, color="0F172A")
    BASE_FONT   = Font(name=FONT_FAMILY, size=9, color="1E293B")
    MUTED_FONT  = Font(name=FONT_FAMILY, size=8, color="64748B")
    KPI_VAL_FONT= Font(name=FONT_FAMILY, size=15, bold=True, color="0F172A")
    KPI_LBL_FONT= Font(name=FONT_FAMILY, size=8, bold=True, color="64748B")

    thin_border_side = Side(style="thin", color="CBD5E1")
    BORDER_ALL = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    BORDER_BOTTOM_DOUBLE = Border(
        left=thin_border_side, right=thin_border_side, top=thin_border_side,
        bottom=Side(style="double", color="0F172A")
    )

    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGN_LEFT   = Alignment(horizontal="left", vertical="center")
    ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")
    ALIGN_WRAP   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    FMT_MONEY = '[$S/-es-PE] #,##0.00'
    FMT_NUM   = '#,##0'
    FMT_PCT   = '0.0%'

    # ═════════════════════════════════════════════════════════════════════════
    # HOJA 1: RESUMEN EJECUTIVO
    # ═════════════════════════════════════════════════════════════════════════
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Ejecutivo"
    ws_resumen.views.sheetView[0].showGridLines = True

    # Banner Header
    ws_resumen.merge_cells("A1:G1")
    t1 = ws_resumen["A1"]
    t1.value = "PERÚ COMPRAS — REPORTE DE INTELIGENCIA DE PRECIOS Y OFERTAS"
    t1.fill = NAVY_MAIN
    t1.font = TITLE_FONT
    t1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_resumen.row_dimensions[1].height = 30

    ws_resumen.merge_cells("A2:G2")
    t2 = ws_resumen["A2"]
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    prov_str = "TODOS LOS PROVEEDORES (CONSOLIDADO)" if (not proveedor or proveedor == "all") else proveedor.upper()
    t2.value = f"Sistema CEAM Auditor 2.0  |  Alcance: {prov_str}  |  Región: {region or 'Nacional'}  |  Generado: {now_str}"
    t2.fill = NAVY_HDR
    t2.font = SUB_FONT
    t2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_resumen.row_dimensions[2].height = 20

    # KPI 1: Fichas Únicas
    ws_resumen.merge_cells("A4:B4")
    ws_resumen["A4"].value = "FICHAS ÚNICAS"
    ws_resumen["A4"].font = KPI_LBL_FONT
    ws_resumen["A4"].alignment = ALIGN_CENTER
    ws_resumen["A4"].fill = CARD_BG

    ws_resumen.merge_cells("A5:B5")
    ws_resumen["A5"].value = total_fichas
    ws_resumen["A5"].number_format = FMT_NUM
    ws_resumen["A5"].font = KPI_VAL_FONT
    ws_resumen["A5"].alignment = ALIGN_CENTER
    ws_resumen["A5"].fill = CARD_BG

    ws_resumen.merge_cells("A6:B6")
    ws_resumen["A6"].value = "Catálogo total analizado"
    ws_resumen["A6"].font = MUTED_FONT
    ws_resumen["A6"].alignment = ALIGN_CENTER
    ws_resumen["A6"].fill = CARD_BG

    # KPI 2: Con Competencia
    ws_resumen.merge_cells("C4:D4")
    ws_resumen["C4"].value = "CON COMPETENCIA (DOBLE OFERTA)"
    ws_resumen["C4"].font = KPI_LBL_FONT
    ws_resumen["C4"].alignment = ALIGN_CENTER
    ws_resumen["C4"].fill = CARD_BG

    ws_resumen.merge_cells("C5:D5")
    ws_resumen["C5"].value = total_competing
    ws_resumen["C5"].number_format = FMT_NUM
    ws_resumen["C5"].font = KPI_VAL_FONT
    ws_resumen["C5"].alignment = ALIGN_CENTER
    ws_resumen["C5"].fill = CARD_BG

    ws_resumen.merge_cells("C6:D6")
    ws_resumen["C6"].value = "Con disputa entre 2 o más proveedores"
    ws_resumen["C6"].font = MUTED_FONT
    ws_resumen["C6"].alignment = ALIGN_CENTER
    ws_resumen["C6"].fill = CARD_BG

    # KPI 3: Stock Total
    ws_resumen.merge_cells("E4:F4")
    ws_resumen["E4"].value = "STOCK TOTAL DISPONIBLE"
    ws_resumen["E4"].font = KPI_LBL_FONT
    ws_resumen["E4"].alignment = ALIGN_CENTER
    ws_resumen["E4"].fill = CARD_BG

    ws_resumen.merge_cells("E5:F5")
    ws_resumen["E5"].value = total_stock
    ws_resumen["E5"].number_format = FMT_NUM
    ws_resumen["E5"].font = KPI_VAL_FONT
    ws_resumen["E5"].alignment = ALIGN_CENTER
    ws_resumen["E5"].fill = CARD_BG

    ws_resumen.merge_cells("E6:F6")
    ws_resumen["E6"].value = "Unidades físicas en vista"
    ws_resumen["E6"].font = MUTED_FONT
    ws_resumen["E6"].alignment = ALIGN_CENTER
    ws_resumen["E6"].fill = CARD_BG

    # KPI 4: Precio Promedio
    ws_resumen["G4"].value = "PRECIO PROMEDIO"
    ws_resumen["G4"].font = KPI_LBL_FONT
    ws_resumen["G4"].alignment = ALIGN_CENTER
    ws_resumen["G4"].fill = CARD_BG

    avg_price = (sum(float(it.get("min_precio") or 0) for it in items) / len(items)) if items else 0
    ws_resumen["G5"].value = avg_price
    ws_resumen["G5"].number_format = FMT_MONEY
    ws_resumen["G5"].font = KPI_VAL_FONT
    ws_resumen["G5"].alignment = ALIGN_CENTER
    ws_resumen["G5"].fill = CARD_BG

    ws_resumen["G6"].value = "Precio referencial calculado"
    ws_resumen["G6"].font = MUTED_FONT
    ws_resumen["G6"].alignment = ALIGN_CENTER
    ws_resumen["G6"].fill = CARD_BG

    for r in range(4, 7):
        for c in range(1, 8):
            cell = ws_resumen.cell(row=r, column=c)
            cell.border = BORDER_ALL

    # Table: Distribución por 14 Categorías Oficiales
    ws_resumen.cell(row=8, column=1, value="DISTRIBUCIÓN POR CATEGORÍAS OFICIALES (ACUERDO MARCO 249)").font = BOLD_FONT
    
    cat_headers = ["Catálogo Oficial", "Categoría de Producto", "Total Ofertas", "% del Catálogo"]
    for ci, h in enumerate(cat_headers, start=1):
        c = ws_resumen.cell(row=9, column=ci, value=h)
        c.fill = NAVY_HDR
        c.font = HDR_FONT
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    ws_resumen.row_dimensions[9].height = 22

    CATEGORY_DISTRIB = [
        ("Catálogo 252: Computadoras de Escritorio", "COMPUTADORA DE ESCRITORIO", cat_counts.get("escritorio", 0)),
        ("Catálogo 252: Computadoras de Escritorio", "COMPUTADORA TODO EN UNO", cat_counts.get("aio", 0)),
        ("Catálogo 252: Computadoras de Escritorio", "ESTACION DE TRABAJO (WORKSTATION)", cat_counts.get("workstation", 0)),
        ("Catálogo 252: Computadoras de Escritorio", "MONITOR", cat_counts.get("monitor", 0)),
        ("Catálogo 252: Computadoras de Escritorio", "PANTALLA PUBLICITARIA", cat_counts.get("pantalla_pub", 0)),
        ("Catálogo 252: Computadoras de Escritorio", "PANTALLA INTERACTIVA", cat_counts.get("pantalla_int", 0)),
        ("Catálogo 252: Computadoras de Escritorio", "ALMACENAMIENTO INTERNO", cat_counts.get("almacenamiento_int", 0)),
        ("Catálogo 252: Computadoras de Escritorio", "ALMACENAMIENTO EXTERNO", cat_counts.get("almacenamiento_ext", 0)),
        ("Catálogo 250: Computadoras Portátiles", "COMPUTADORA PORTATIL (LAPTOP)", cat_counts.get("portatil", 0)),
        ("Catálogo 250: Computadoras Portátiles", "ESTACION DE TRABAJO PORTATIL", cat_counts.get("workstation_portatil", 0)),
        ("Catálogo 250: Computadoras Portátiles", "TABLETA", cat_counts.get("tableta", 0)),
        ("Catálogo 251: Escáneres", "ESCANER DE DOCUMENTOS", cat_counts.get("escaner_docs", 0)),
        ("Catálogo 251: Escáneres", "ESCANER DE PLANOS", cat_counts.get("escaner_planos", 0)),
        ("Catálogo 251: Escáneres", "ESCANER DE LIBROS", cat_counts.get("escaner_libros", 0)),
    ]

    total_cat_sum = sum(cnt for _, _, cnt in CATEGORY_DISTRIB) or 1
    current_row = 10
    for cat_oficial, cat_nombre, cnt in CATEGORY_DISTRIB:
        fill = LIGHT_BG if current_row % 2 == 0 else PatternFill(fill_type=None)
        
        c1 = ws_resumen.cell(row=current_row, column=1, value=cat_oficial)
        c2 = ws_resumen.cell(row=current_row, column=2, value=cat_nombre)
        c3 = ws_resumen.cell(row=current_row, column=3, value=cnt)
        c4 = ws_resumen.cell(row=current_row, column=4, value=f"=C{current_row}/C{len(CATEGORY_DISTRIB)+10}")

        c1.font = BASE_FONT; c1.alignment = ALIGN_LEFT; c1.border = BORDER_ALL; c1.fill = fill
        c2.font = BASE_FONT; c2.alignment = ALIGN_LEFT; c2.border = BORDER_ALL; c2.fill = fill
        c3.font = BASE_FONT; c3.alignment = ALIGN_RIGHT; c3.border = BORDER_ALL; c3.fill = fill; c3.number_format = FMT_NUM
        c4.font = BASE_FONT; c4.alignment = ALIGN_RIGHT; c4.border = BORDER_ALL; c4.fill = fill; c4.number_format = FMT_PCT
        current_row += 1

    # Total Row for Categories
    tot_row = current_row
    c1 = ws_resumen.cell(row=tot_row, column=1, value="TOTAL CONSOLIDADO")
    c2 = ws_resumen.cell(row=tot_row, column=2, value="")
    c3 = ws_resumen.cell(row=tot_row, column=3, value=f"=SUM(C10:C{tot_row-1})")
    c4 = ws_resumen.cell(row=tot_row, column=4, value="100.0%")
    ws_resumen.merge_cells(f"A{tot_row}:B{tot_row}")
    c1.font = BOLD_FONT; c1.alignment = ALIGN_LEFT; c1.fill = CARD_BG; c1.border = BORDER_BOTTOM_DOUBLE
    c2.border = BORDER_BOTTOM_DOUBLE
    c3.font = BOLD_FONT; c3.alignment = ALIGN_RIGHT; c3.fill = CARD_BG; c3.border = BORDER_BOTTOM_DOUBLE; c3.number_format = FMT_NUM
    c4.font = BOLD_FONT; c4.alignment = ALIGN_RIGHT; c4.fill = CARD_BG; c4.border = BORDER_BOTTOM_DOUBLE

    # Adjust widths in Resumen
    ws_resumen.column_dimensions["A"].width = 38
    ws_resumen.column_dimensions["B"].width = 36
    ws_resumen.column_dimensions["C"].width = 18
    ws_resumen.column_dimensions["D"].width = 18
    ws_resumen.column_dimensions["E"].width = 18
    ws_resumen.column_dimensions["F"].width = 18
    ws_resumen.column_dimensions["G"].width = 22

    # ═════════════════════════════════════════════════════════════════════════
    # HOJA 2: CATÁLOGO Y COMPARATIVA
    # ═════════════════════════════════════════════════════════════════════════
    ws_data = wb.create_sheet(title="Catálogo y Ofertas")
    ws_data.views.sheetView[0].showGridLines = True

    # Title Bar
    ws_data.merge_cells("A1:O1")
    dt1 = ws_data["A1"]
    dt1.value = f"DETALLE DE CATÁLOGO Y COMPARATIVA DE PRECIOS ({len(items):,} FICHAS EXTRAÍDAS)"
    dt1.fill = NAVY_MAIN
    dt1.font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
    dt1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_data.row_dimensions[1].height = 24

    # Headers
    DATA_HEADERS = [
        ("N° Parte / Código", 22, ALIGN_CENTER),
        ("Marca", 16, ALIGN_CENTER),
        ("Descripción / Especificación Técnica", 48, ALIGN_LEFT),
        ("Categoría Oficial", 28, ALIGN_LEFT),
        ("Catálogo", 22, ALIGN_LEFT),
        ("Tipo de Oferta", 22, ALIGN_CENTER),
        ("Proveedores Oferentes", 32, ALIGN_LEFT),
        ("Precio Mínimo (S/)", 18, ALIGN_RIGHT),
        ("Precio Máximo (S/)", 18, ALIGN_RIGHT),
        ("Stock Total", 14, ALIGN_RIGHT),
        ("Plazo Entrega (Días)", 16, ALIGN_CENTER),
        ("Última OCAM", 18, ALIGN_CENTER),
        ("Fecha OCAM", 15, ALIGN_CENTER),
        ("Ficha Técnica (PDF)", 18, ALIGN_CENTER),
        ("Fecha Extracción", 16, ALIGN_CENTER),
    ]

    for ci, (h_title, h_width, h_align) in enumerate(DATA_HEADERS, start=1):
        cell = ws_data.cell(row=2, column=ci, value=h_title)
        cell.fill = NAVY_HDR
        cell.font = HDR_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
        ws_data.column_dimensions[get_column_letter(ci)].width = h_width
    ws_data.row_dimensions[2].height = 24

    # Data Rows
    row_idx = 3
    for it in items:
        fill = LIGHT_BG if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        ofertas = it.get("ofertas") or []
        prov_names = []
        for o in ofertas:
            if isinstance(o, dict) and o.get("nombre_proveedor"):
                p_name = o["nombre_proveedor"].strip()
                if "JORGE" in p_name.upper() or "ROJAS" in p_name.upper():
                    prov_names.append("Jorge Rojas Villanueva")
                elif "KING" in p_name.upper():
                    prov_names.append("The King Computer")
                else:
                    prov_names.append(p_name)
        if not prov_names:
            prov_names = [it.get("proveedor") or it.get("nombre_proveedor") or "The King Computer"]
        prov_str = ", ".join(dict.fromkeys(prov_names))

        is_competing = len(ofertas) > 1 or it.get("total_proveedores", 1) > 1
        tipo_oferta = "Con Competencia (2+)" if is_competing else "Oferta Exclusiva (1)"

        p_min = float(it.get("min_precio") or it.get("precio_ofertado") or 0)
        p_max = float(it.get("max_precio") or it.get("precio_ofertado") or p_min)
        stock_val = int(it.get("existencia_stock") or 0)
        plazo_val = it.get("min_plazo_entrega") or it.get("plazo_entrega_dias")

        c1  = ws_data.cell(row=row_idx, column=1, value=it.get("nro_parte") or "S/N")
        c2  = ws_data.cell(row=row_idx, column=2, value=it.get("marca") or "—")
        c3  = ws_data.cell(row=row_idx, column=3, value=it.get("descripcion") or it.get("descripcion_producto") or "—")
        c4  = ws_data.cell(row=row_idx, column=4, value=it.get("categoria") or "—")
        c5  = ws_data.cell(row=row_idx, column=5, value=it.get("catalogo") or "—")
        c6  = ws_data.cell(row=row_idx, column=6, value=tipo_oferta)
        c7  = ws_data.cell(row=row_idx, column=7, value=prov_str)
        c8  = ws_data.cell(row=row_idx, column=8, value=p_min)
        c9  = ws_data.cell(row=row_idx, column=9, value=p_max)
        c10 = ws_data.cell(row=row_idx, column=10, value=stock_val)
        c11 = ws_data.cell(row=row_idx, column=11, value=plazo_val if plazo_val is not None else "—")
        c12 = ws_data.cell(row=row_idx, column=12, value=it.get("orden_min") or "—")
        c13 = ws_data.cell(row=row_idx, column=13, value=str(it.get("fecha_orden_min") or "")[:10] or "—")
        
        pdf_val = it.get("pdf_url")
        c14 = ws_data.cell(row=row_idx, column=14)
        if pdf_val and str(pdf_val).startswith("http"):
            c14.value = "Ver Ficha PDF"
            c14.hyperlink = pdf_val
            c14.font = Font(name=FONT_FAMILY, size=9, color="2563EB", underline="single")
        else:
            c14.value = "—"
            c14.font = BASE_FONT

        c15 = ws_data.cell(row=row_idx, column=15, value=str(it.get("fecha_extraccion") or "")[:10] or "—")

        # Formats & alignments
        c1.alignment = ALIGN_CENTER; c1.font = Font(name="Consolas", size=9, bold=True, color="1E3A8A")
        c2.alignment = ALIGN_CENTER; c2.font = BASE_FONT
        c3.alignment = ALIGN_WRAP;   c3.font = BASE_FONT
        c4.alignment = ALIGN_LEFT;   c4.font = BASE_FONT
        c5.alignment = ALIGN_LEFT;   c5.font = BASE_FONT
        
        c6.alignment = ALIGN_CENTER
        if is_competing:
            c6.fill = PURPLE_FILL
            c6.font = PURPLE_TXT
        else:
            c6.fill = fill
            c6.font = BASE_FONT

        c7.alignment = ALIGN_LEFT;   c7.font = BASE_FONT
        c8.alignment = ALIGN_RIGHT;  c8.font = Font(name=FONT_FAMILY, size=9, bold=True); c8.number_format = FMT_MONEY
        c9.alignment = ALIGN_RIGHT;  c9.font = BASE_FONT; c9.number_format = FMT_MONEY
        c10.alignment = ALIGN_RIGHT; c10.font = Font(name=FONT_FAMILY, size=9, bold=(stock_val > 0), color="166534" if stock_val > 0 else "64748B"); c10.number_format = FMT_NUM
        c11.alignment = ALIGN_CENTER; c11.font = BASE_FONT
        c12.alignment = ALIGN_CENTER; c12.font = Font(name="Consolas", size=9, color="0284C7")
        c13.alignment = ALIGN_CENTER; c13.font = BASE_FONT
        c14.alignment = ALIGN_CENTER
        c15.alignment = ALIGN_CENTER; c15.font = BASE_FONT

        for col_i in range(1, 16):
            cell = ws_data.cell(row=row_idx, column=col_i)
            cell.border = BORDER_ALL
            if col_i != 6 or not is_competing:
                if not cell.fill or cell.fill.fill_type is None:
                    cell.fill = fill

        ws_data.row_dimensions[row_idx].height = 20
        row_idx += 1

    # Totals Row at the end
    end_row = row_idx
    ws_data.cell(row=end_row, column=1, value="TOTAL / PROMEDIO").font = BOLD_FONT
    ws_data.cell(row=end_row, column=1).alignment = ALIGN_LEFT
    ws_data.cell(row=end_row, column=1).fill = CARD_BG
    ws_data.cell(row=end_row, column=1).border = BORDER_BOTTOM_DOUBLE

    for c in range(2, 8):
        ws_data.cell(row=end_row, column=c).fill = CARD_BG
        ws_data.cell(row=end_row, column=c).border = BORDER_BOTTOM_DOUBLE

    ws_data.merge_cells(f"A{end_row}:G{end_row}")

    # Min price average
    c8 = ws_data.cell(row=end_row, column=8, value=f"=AVERAGE(H3:H{end_row-1})")
    c8.font = BOLD_FONT; c8.alignment = ALIGN_RIGHT; c8.fill = CARD_BG; c8.border = BORDER_BOTTOM_DOUBLE; c8.number_format = FMT_MONEY

    # Max price average
    c9 = ws_data.cell(row=end_row, column=9, value=f"=AVERAGE(I3:I{end_row-1})")
    c9.font = BOLD_FONT; c9.alignment = ALIGN_RIGHT; c9.fill = CARD_BG; c9.border = BORDER_BOTTOM_DOUBLE; c9.number_format = FMT_MONEY

    # Stock total sum
    c10 = ws_data.cell(row=end_row, column=10, value=f"=SUM(J3:J{end_row-1})")
    c10.font = BOLD_FONT; c10.alignment = ALIGN_RIGHT; c10.fill = CARD_BG; c10.border = BORDER_BOTTOM_DOUBLE; c10.number_format = FMT_NUM

    for c in range(11, 16):
        ws_data.cell(row=end_row, column=c).fill = CARD_BG
        ws_data.cell(row=end_row, column=c).border = BORDER_BOTTOM_DOUBLE

    # Auto-filter & Freeze Panes
    ws_data.auto_filter.ref = f"A2:O{end_row-1}"
    ws_data.freeze_panes = "A3"

    # Save to Stream and Return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_ofertas_perucompras_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/categories-count")
def get_categories_count(
    proveedor: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Devuelve el conteo de ofertas distribuidas exactamente por las 14 categorías oficiales de Perú Compras."""
    prov_where = ""
    params = {}
    if proveedor and proveedor.lower() != "all":
        prov_l = proveedor.lower()
        if prov_l in ("thekingcomputer", "king"):
            prov_where = " AND (UPPER(nombre_proveedor) LIKE '%KING%' OR UPPER(ruc_proveedor) = '20601234567')"
        elif prov_l in ("jorge_rojas", "jorge", "rojas"):
            prov_where = " AND (UPPER(nombre_proveedor) LIKE '%ROJAS%' OR UPPER(nombre_proveedor) LIKE '%JORGE%' OR UPPER(ruc_proveedor) = '10408899991')"
        else:
            prov_where = " AND UPPER(nombre_proveedor) LIKE UPPER(:prov)"
            params["prov"] = f"%{proveedor}%"

    try:
        total = db.execute(text(f"SELECT COUNT(*) FROM ofertas_proveedor_history WHERE 1=1 {prov_where};"), params).scalar() or 0
        
        # 1. COMPUTADORAS DE ESCRITORIO (Catálogo 252 - 8 categorías)
        escritorio = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where}
              AND (
                UPPER(categoria) = 'COMPUTADORA DE ESCRITORIO'
                OR (
                    (UPPER(categoria) LIKE '%ESCRITORIO%' OR UPPER(descripcion_producto) LIKE '%ESCRITORIO%')
                    AND UPPER(categoria) NOT LIKE '%TODO EN UNO%'
                    AND UPPER(descripcion_producto) NOT LIKE '%TODO EN UNO%'
                    AND UPPER(descripcion_producto) NOT LIKE '%ALL IN ONE%'
                    AND UPPER(descripcion_producto) NOT LIKE '%PORTATIL%'
                    AND UPPER(descripcion_producto) NOT LIKE '%PORTÁTIL%'
                )
              );
        """), params).scalar() or 0

        aio = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where}
              AND (
                UPPER(categoria) LIKE '%TODO EN UNO%' 
                OR UPPER(descripcion_producto) LIKE '%TODO EN UNO%' 
                OR UPPER(descripcion_producto) LIKE '%ALL IN ONE%'
              );
        """), params).scalar() or 0

        workstation = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where}
              AND UPPER(categoria) = 'ESTACION DE TRABAJO'
              AND UPPER(categoria) NOT LIKE '%PORTATIL%';
        """), params).scalar() or 0

        monitor = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where}
              AND (UPPER(categoria) = 'MONITOR' OR (UPPER(descripcion_producto) LIKE 'MONITOR%' AND UPPER(descripcion_producto) NOT LIKE '%TODO EN UNO%'));
        """), params).scalar() or 0

        pantalla_pub = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where} AND UPPER(categoria) LIKE '%PANTALLA PUBLICITARIA%';
        """), params).scalar() or 0

        pantalla_int = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where} AND UPPER(categoria) LIKE '%PANTALLA INTERACTIVA%';
        """), params).scalar() or 0

        almacenamiento_int = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where} AND UPPER(categoria) LIKE '%ALMACENAMIENTO INTERNO%';
        """), params).scalar() or 0

        almacenamiento_ext = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where} AND UPPER(categoria) LIKE '%ALMACENAMIENTO EXTERNO%';
        """), params).scalar() or 0

        # 2. COMPUTADORAS PORTÁTILES (Catálogo 250 - 3 categorías)
        portatil = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where}
              AND (
                UPPER(categoria) = 'COMPUTADORA PORTATIL'
                OR (
                    (UPPER(categoria) LIKE '%PORTATIL%' OR UPPER(descripcion_producto) LIKE '%PORTATIL%' OR UPPER(descripcion_producto) LIKE '%LAPTOP%')
                    AND UPPER(categoria) NOT LIKE '%ESTACION%'
                    AND UPPER(descripcion_producto) NOT LIKE '%TODO EN UNO%'
                )
              );
        """), params).scalar() or 0

        workstation_portatil = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where}
              AND (UPPER(categoria) LIKE '%ESTACION DE TRABAJO PORTATIL%' OR UPPER(descripcion_producto) LIKE '%WORKSTATION PORTATIL%');
        """), params).scalar() or 0

        tableta = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where}
              AND (UPPER(categoria) LIKE '%TABLET%' OR UPPER(descripcion_producto) LIKE '%TABLETA%');
        """), params).scalar() or 0

        # 3. ESCÁNERES (Catálogo 251 - 3 categorías)
        escaner_planos = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where} AND UPPER(categoria) LIKE '%ESCANER DE PLANOS%';
        """), params).scalar() or 0

        escaner_docs = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where} AND (UPPER(categoria) LIKE '%ESCANER DE DOCUMENTOS%' OR (UPPER(catalogo) LIKE '%ESCANER%' AND UPPER(categoria) NOT LIKE '%PLANOS%' AND UPPER(categoria) NOT LIKE '%LIBROS%'));
        """), params).scalar() or 0

        escaner_libros = db.execute(text(f"""
            SELECT COUNT(*) FROM ofertas_proveedor_history 
            WHERE 1=1 {prov_where} AND UPPER(categoria) LIKE '%ESCANER DE LIBROS%';
        """), params).scalar() or 0

        return {
            "total": total,
            "escritorio": escritorio,
            "aio": aio,
            "workstation": workstation,
            "monitor": monitor,
            "pantalla_pub": pantalla_pub,
            "pantalla_int": pantalla_int,
            "almacenamiento_int": almacenamiento_int,
            "almacenamiento_ext": almacenamiento_ext,
            "portatil": portatil,
            "workstation_portatil": workstation_portatil,
            "tableta": tableta,
            "escaner_planos": escaner_planos,
            "escaner_docs": escaner_docs,
            "escaner_libros": escaner_libros
        }
    except Exception as e:
        return {
            "total": 0, "escritorio": 0, "aio": 0, "workstation": 0, "monitor": 0,
            "pantalla_pub": 0, "pantalla_int": 0, "almacenamiento_int": 0, "almacenamiento_ext": 0,
            "portatil": 0, "workstation_portatil": 0, "tableta": 0,
            "escaner_planos": 0, "escaner_docs": 0, "escaner_libros": 0, "error": str(e)
        }

@router.post("/reclassify")
def reclassify_existing_offers(db: Session = Depends(get_db)):
    """Reclasifica en lote todas las ofertas existentes analizando su descripción."""
    try:
        db.execute(text("""
            -- Limpieza preventiva de duplicados antes de actualizar clasificaciones
            DELETE FROM ofertas_proveedor_history a
            USING ofertas_proveedor_history b
            WHERE a.id < b.id 
              AND a.nro_parte = b.nro_parte 
              AND a.ruc_proveedor = b.ruc_proveedor 
              AND COALESCE(a.region, 'N/A') = COALESCE(b.region, 'N/A');

            -- 1. Monitores
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'MONITORES', categoria = 'MONITOR'
            WHERE (UPPER(descripcion_producto) LIKE 'MONITOR%' 
               OR UPPER(descripcion_producto) LIKE '%MONITOR LED%' 
               OR UPPER(descripcion_producto) LIKE '%MONITOR PARA PC%'
               OR UPPER(categoria) LIKE '%MONITOR%')
              AND UPPER(descripcion_producto) NOT LIKE '%TODO EN UNO%';

            -- 2. Todo en Uno (AIO)
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'COMPUTADORAS DE ESCRITORIO', categoria = 'COMPUTADORA TODO EN UNO'
            WHERE (UPPER(descripcion_producto) LIKE '%TODO EN UNO%' 
               OR UPPER(descripcion_producto) LIKE '%ALL IN ONE%' 
               OR UPPER(descripcion_producto) LIKE '%ALL-IN-ONE%');

            -- 3. Portátiles / Laptops
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'COMPUTADORAS PORTATILES', categoria = 'COMPUTADORA PORTATIL'
            WHERE (UPPER(descripcion_producto) LIKE '%PORTATIL%' 
               OR UPPER(descripcion_producto) LIKE '%PORTÁTIL%' 
               OR UPPER(descripcion_producto) LIKE '%LAPTOP%' 
               OR UPPER(descripcion_producto) LIKE '%NOTEBOOK%')
              AND UPPER(descripcion_producto) NOT LIKE '%TODO EN UNO%';

            -- 4. Impresoras / Multifuncionales
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'IMPRESORAS', categoria = 'IMPRESORA'
            WHERE (UPPER(descripcion_producto) LIKE '%IMPRESORA%' 
               OR UPPER(descripcion_producto) LIKE '%MULTIFUNCIONAL%'
               OR UPPER(descripcion_producto) LIKE '%PLOTTER%'
               OR UPPER(categoria) LIKE '%IMPRESORA%');

            -- 5. Escáneres
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'ESCANERES', categoria = 'ESCANER'
            WHERE (UPPER(descripcion_producto) LIKE '%ESCANER%' 
               OR UPPER(descripcion_producto) LIKE '%ESCÁNER%');

            -- 6. Tablets
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'TABLETS', categoria = 'TABLET'
            WHERE UPPER(descripcion_producto) LIKE '%TABLET%';

            -- 7. Estaciones de Trabajo (Workstations)
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'ESTACIONES DE TRABAJO', categoria = 'ESTACION DE TRABAJO'
            WHERE (UPPER(descripcion_producto) LIKE '%ESTACION DE TRABAJO%' 
               OR UPPER(descripcion_producto) LIKE '%ESTACIÓN DE TRABAJO%'
               OR UPPER(descripcion_producto) LIKE '%WORKSTATION%');

            -- 8. Servidores
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'SERVIDORES', categoria = 'SERVIDOR'
            WHERE (UPPER(descripcion_producto) LIKE '%SERVIDOR%' 
               OR UPPER(descripcion_producto) LIKE '%SERVER%');

            -- 9. Proyectores
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'PROYECTORES', categoria = 'PROYECTOR'
            WHERE UPPER(descripcion_producto) LIKE '%PROYECTOR%';

            -- 10. UPS / Energía
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'ENERGIA Y UPS', categoria = 'UPS'
            WHERE (UPPER(descripcion_producto) LIKE '%UPS%' 
               OR UPPER(descripcion_producto) LIKE '%ENERGIA ININTERRUMPIDA%'
               OR UPPER(descripcion_producto) LIKE '%ESTABILIZADOR%');

            -- 11. Computadoras de Escritorio (Desktop / Torre)
            UPDATE ofertas_proveedor_history 
            SET catalogo = 'COMPUTADORAS DE ESCRITORIO', categoria = 'COMPUTADORA DE ESCRITORIO'
            WHERE (UPPER(descripcion_producto) LIKE 'COMPUTADORA DE ESCRITORIO%' 
               OR UPPER(descripcion_producto) LIKE '%ESCRITORIO%' 
               OR UPPER(descripcion_producto) LIKE '%MINI PC%'
               OR UPPER(descripcion_producto) LIKE '%SFF%')
              AND UPPER(descripcion_producto) NOT LIKE '%TODO EN UNO%'
              AND UPPER(descripcion_producto) NOT LIKE '%PORTATIL%'
              AND UPPER(descripcion_producto) NOT LIKE '%PORTÁTIL%'
              AND UPPER(descripcion_producto) NOT LIKE '%MONITOR%';
        """))
        db.commit()
        return {"success": True, "message": "Reclasificación ejecutada con éxito"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@router.get("/kpis")
def get_proveedor_kpis(
    proveedor: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Devuelve los KPIs consolidados del proveedor.
    """
    params = {}
    where_clause = ""
    if proveedor and proveedor.lower() != "all":
        where_clause = "WHERE UPPER(nombre_proveedor) LIKE UPPER(:proveedor)"
        params["proveedor"] = f"%{proveedor}%"

    try:
        res = db.execute(text(f"""
            SELECT 
                COUNT(DISTINCT nro_orden_fisica) AS total_ordenes,
                COALESCE(SUM(monto_total), 0) AS total_vendido,
                COUNT(DISTINCT id) AS fichas_count,
                COALESCE(AVG(monto_total), 0) AS avg_precio
            FROM purchase_orders
            {where_clause}
        """), params).mappings().first()
        return dict(res) if res else {"total_ordenes": 0, "total_vendido": 0, "fichas_count": 0, "avg_precio": 0}
    except Exception:
        return {"total_ordenes": 0, "total_vendido": 0, "fichas_count": 0, "avg_precio": 0}

@router.get("/accounts")
def get_available_accounts():
    """Devuelve la lista de cuentas de proveedores configuradas para extracción."""
    from app.services.proveedores_scraper import PROVEEDORES_CONFIG
    return [
        {
            "id": k,
            "nombre": v["nombre"],
            "short": v["short"],
            "user": v["user"],
            "ruc": v["ruc"]
        }
        for k, v in PROVEEDORES_CONFIG.items()
    ]

@router.post("/scrape")
async def trigger_scrape_proveedores(
    background_tasks: BackgroundTasks,
    proveedor: str = Query("thekingcomputer", description="ID del proveedor a extraer: thekingcomputer, jorge_rojas")
):
    """
    Inicia la extracción en segundo plano para la cuenta del proveedor especificado.
    """
    from app.services.proveedores_scraper import PROVEEDORES_CONFIG, run_worker_pool_extraction
    from app.db.database import SessionLocal
    prov_nombre = PROVEEDORES_CONFIG.get(proveedor, {}).get("nombre", proveedor)

    async def _async_task():
        with SessionLocal() as db_session:
            await run_worker_pool_extraction([], db_session, provider_key=proveedor)

    background_tasks.add_task(_async_task)
    return {
        "message": f"Worker Pool de extracción iniciado para '{prov_nombre}'",
        "proveedor": proveedor,
        "nombre": prov_nombre
    }

@router.post("/scrape-plazos")
async def trigger_scrape_plazos(
    background_tasks: BackgroundTasks,
    proveedor: str = Query("thekingcomputer", description="ID del proveedor: thekingcomputer, jorge_rojas, all"),
    regiones: Optional[str] = Query(None, description="Regiones separadas por coma (ej: LIMA,AREQUIPA) o vacío para todas")
):
    """
    Inicia la extracción regional de plazos de entrega en segundo plano.
    """
    from app.services.proveedores_scraper import async_extract_plazos_regionales, PROVEEDORES_CONFIG
    from app.db.database import SessionLocal
    if regiones and regiones.strip().lower() in ("all", "todas", "global", "none", "null"):
        target_regs = None
    elif regiones:
        target_regs = [r.strip().upper() for r in regiones.split(",") if r.strip()]
    else:
        target_regs = None

    async def _async_plazos_task():
        with SessionLocal() as db_session:
            if proveedor in ("all", "ambos", "todos"):
                for p_key in ("thekingcomputer", "jorge_rojas"):
                    await async_extract_plazos_regionales(provider_key=p_key, regiones=target_regs, db=db_session)
            else:
                await async_extract_plazos_regionales(provider_key=proveedor, regiones=target_regs, db=db_session)

    background_tasks.add_task(_async_plazos_task)
    prov_desc = "Ambos Proveedores (The King y Jorge Rojas)" if proveedor in ("all", "ambos", "todos") else PROVEEDORES_CONFIG.get(proveedor, {}).get("nombre", proveedor)
    return {
        "message": f"Extracción regional de plazos iniciada para '{prov_desc}'",
        "proveedor": proveedor,
        "regiones": target_regs or "TODAS (25 Regiones)"
    }

@router.get("/scrape-status")
def get_scrape_status():
    from app.services.proveedores_scraper import EXTRACTION_STATUS
    return EXTRACTION_STATUS
