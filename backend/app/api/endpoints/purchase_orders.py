"""Purchase Orders REST endpoints."""
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session

from app.db.database import get_db
from app.schemas.purchase_order import PurchaseOrderCreate, PurchaseOrderResponse
from app.services import crud

router = APIRouter(prefix="/purchase-orders", tags=["Purchase Orders"])


@router.get("/", response_model=List[PurchaseOrderResponse])
def list_orders(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    catalogo: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    estado_orden: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    entidad: Optional[str] = Query(None),
    proveedor: Optional[str] = Query(None),
    marca: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    sort_dir: Optional[str] = Query("desc"),
    db: Session = Depends(get_db),
):
    """List purchase orders with optional filters, sorting, and pagination."""
    return crud.get_orders(
        db,
        skip=skip,
        limit=limit,
        catalogo=catalogo,
        categoria=categoria,
        estado_orden=estado_orden,
        search=search,
        entidad=entidad,
        proveedor=proveedor,
        marca=marca,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )


@router.get("/summary")
def get_orders_summary(
    catalogo: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    estado_orden: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    entidad: Optional[str] = Query(None),
    proveedor: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Return total count of purchase orders matching the provided filters."""
    total = crud.count_orders_filtered(
        db,
        catalogo=catalogo,
        categoria=categoria,
        estado_orden=estado_orden,
        search=search,
        entidad=entidad,
        proveedor=proveedor,
    )
    return {"total": total}


@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    """Aggregated statistics for dashboard KPIs and charts."""
    return crud.get_stats(db)


@router.get("/catalogos-filter")
def get_catalogos_filter(db: Session = Depends(get_db)):
    """Return distinct catalogo values present in the DB for the filter dropdown."""
    from app.models.purchase_order import PurchaseOrder
    rows = (
        db.query(PurchaseOrder.catalogo)
        .filter(PurchaseOrder.catalogo.isnot(None))
        .distinct()
        .order_by(PurchaseOrder.catalogo)
        .all()
    )
    return {"catalogos": [r[0] for r in rows if r[0]]}


@router.get("/filters/{column_name}")
def get_column_filters(column_name: str, db: Session = Depends(get_db)):
    """Return distinct non-null values for a specific column to build Excel-like filters."""
    from app.models.purchase_order import PurchaseOrder
    valid_columns = {
        "entidad": PurchaseOrder.nombre_entidad,
        "proveedor": PurchaseOrder.nombre_proveedor,
        "estado": PurchaseOrder.estado_orden,
        "marca": PurchaseOrder.marca,
    }
    if column_name not in valid_columns:
        raise HTTPException(status_code=400, detail="Columna no permitida para filtros")
    
    col = valid_columns[column_name]
    rows = db.query(col).filter(col.isnot(None), col != '').distinct().order_by(col).all()
    return {"values": [r[0] for r in rows if r[0]]}


@router.delete("/all", status_code=200)
def delete_all_orders(db: Session = Depends(get_db)):
    """Delete ALL purchase orders from the database. Used to reset data before a fresh scrape."""
    from app.models.purchase_order import PurchaseOrder
    count = db.query(PurchaseOrder).count()
    db.query(PurchaseOrder).delete()
    db.commit()
    return {"deleted": count, "message": f"Se eliminaron {count} órdenes de compra"}


@router.post("/enrich-marcas")
def enrich_marcas(db: Session = Depends(get_db)):
    """
    Populate purchase_orders.marca from multiple sources (priority order):
      1. fichas_producto — exact JOIN by nro_parte (most reliable).
      2. Regex on detalle_producto against a curated brand list.
      3. Prefix heuristic on nro_parte codes (e.g. HP-, DELL-).
    Returns counts of orders updated per source.
    """
    import json
    import re
    from sqlalchemy import text as _text

    # ── 1. Source A: fichas_producto JOIN by nro_parte ─────────────────────
    # purchase_orders.nro_parte stores a JSON array:
    # [{"nro_parte": "HP-XXXXX", ...}, ...]
    # fichas_producto has columns "nro_parte_o_código_único_de_identificación"
    # (or similar) and "marca".
    # Strategy: for each order, expand JSON products, look up each p/n in fichas,
    # take the first marca found.

    updated_fichas = 0
    updated_regex  = 0
    updated_prefix = 0

    # Detect the nro_parte column name in fichas_producto
    try:
        fich_cols = db.execute(
            _text("SELECT column_name FROM information_schema.columns WHERE table_name='fichas_producto'")
        ).fetchall()
        fich_col_names = [r[0] for r in fich_cols]
        nro_col = next(
            (c for c in fich_col_names if c.startswith("nro_parte")),
            None,
        )
        marca_fichas_col = "marca" if "marca" in fich_col_names else None
    except Exception:
        nro_col = None
        marca_fichas_col = None

    if nro_col and marca_fichas_col:
        # Build dict {nro_parte_upper: marca} from fichas_producto
        try:
            fichas_rows = db.execute(
                _text(f'SELECT UPPER(TRIM("{nro_col}")), UPPER(TRIM("{marca_fichas_col}")) FROM fichas_producto WHERE "{marca_fichas_col}" IS NOT NULL AND "{marca_fichas_col}" != \'\'')
            ).fetchall()
            fichas_map = {r[0]: r[1] for r in fichas_rows if r[0]}
        except Exception:
            fichas_map = {}
    else:
        fichas_map = {}

    # ── 2. Known brand list for regex ──────────────────────────────────────
    KNOWN_BRANDS = [
        "HP", "HEWLETT PACKARD", "DELL", "LENOVO", "ASUS", "ACER", "MSI",
        "APPLE", "SAMSUNG", "LG", "SONY", "TOSHIBA", "EPSON", "CANON",
        "BROTHER", "XEROX", "LEXMARK", "INTEL", "AMD", "NVIDIA",
        "GIGABYTE", "BIOSTAR", "ASRock", "CORSAIR", "KINGSTON", "SEAGATE",
        "WESTERN DIGITAL", "WD", "SANDISK", "CRUCIAL", "LOGITECH",
        "MICROSOFT", "BENQ", "VIEWSONIC", "AOC", "PHILIPS",
        "KENYA", "ABAD", "SECURITAS",
    ]
    # Build compiled regex: word boundary around each brand, case-insensitive
    brand_pattern = re.compile(
        r'\b(' + '|'.join(re.escape(b) for b in KNOWN_BRANDS) + r')\b',
        re.IGNORECASE
    )

    # ── 3. Prefix heuristics ───────────────────────────────────────────────
    PREFIX_MAP = {
        "HP-": "HP", "DELL-": "DELL", "LN-": "LENOVO",
        "AX-": "ASUS", "AC-": "ACER", "MS-": "MSI",
    }

    # ── Process each order without a marca ────────────────────────────────
    from app.models.purchase_order import PurchaseOrder as _PO
    orders_to_enrich = db.query(_PO).filter(
        (_PO.marca.is_(None)) | (_PO.marca == '')
    ).all()

    for order in orders_to_enrich:
        marca_found = None

        # Source A: fichas_producto via nro_parte JSON
        if fichas_map and order.nro_parte:
            try:
                prods = json.loads(order.nro_parte)
                if isinstance(prods, list):
                    for p in prods:
                        key = str(p.get("nro_parte", "")).strip().upper()
                        if key and key in fichas_map:
                            marca_found = fichas_map[key]
                            break
            except Exception:
                pass
            if marca_found:
                updated_fichas += 1

        # Source B: regex on detalle_producto
        if not marca_found and order.detalle_producto:
            m = brand_pattern.search(order.detalle_producto)
            if m:
                marca_found = m.group(1).upper()
                updated_regex += 1

        # Source C: prefix on nro_parte codes
        if not marca_found and order.nro_parte:
            try:
                prods = json.loads(order.nro_parte)
                if isinstance(prods, list):
                    for p in prods:
                        pn = str(p.get("nro_parte", "")).strip().upper()
                        for prefix, brand in PREFIX_MAP.items():
                            if pn.startswith(prefix):
                                marca_found = brand
                                break
                        if marca_found:
                            break
            except Exception:
                pass
            if marca_found:
                updated_prefix += 1

        if marca_found:
            order.marca = marca_found

    db.commit()
    total_updated = updated_fichas + updated_regex + updated_prefix
    return {
        "total_updated": total_updated,
        "from_fichas": updated_fichas,
        "from_regex": updated_regex,
        "from_prefix": updated_prefix,
    }


@router.get("/providers")
def list_providers(db: Session = Depends(get_db)):
    """Return all distinct providers with their order count and total amount."""
    from app.models.purchase_order import PurchaseOrder
    from sqlalchemy import func as f
    rows = (
        db.query(
            PurchaseOrder.nombre_proveedor,
            f.count(PurchaseOrder.id).label("orders"),
            f.sum(PurchaseOrder.monto_total).label("total"),
        )
        .group_by(PurchaseOrder.nombre_proveedor)
        .order_by(f.sum(PurchaseOrder.monto_total).desc())
        .all()
    )
    return {
        "providers": [
            {
                "nombre_proveedor": r[0],
                "orders": r[1],
                "total": float(r[2] or 0),
            }
            for r in rows
            if r[0]
        ]
    }


@router.get("/export")
def export_orders_csv(
    proveedor: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Stream orders as CSV. Filter by proveedor if provided."""
    import csv
    import io
    from fastapi.responses import StreamingResponse

    orders = crud.get_orders(db, skip=0, limit=100_000, proveedor=proveedor)

    def generate():
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "id", "orden_electronica", "nro_orden_fisica", "fecha_publicacion",
            "nombre_entidad", "nombre_proveedor", "catalogo", "categoria",
            "estado_orden", "monto_total", "nro_parte",
        ])
        for o in orders:
            writer.writerow([
                o.id, o.orden_electronica, o.nro_orden_fisica,
                o.fecha_publicacion, o.nombre_entidad, o.nombre_proveedor,
                o.catalogo, o.categoria, o.estado_orden, o.monto_total, o.nro_parte,
            ])
        yield buf.getvalue()

    filename = f"ordenes_{proveedor or 'todas'}.csv"
    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export-excel")
def export_orders_excel(
    catalogo: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    estado_orden: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    entidad: Optional[str] = Query(None),
    proveedor: Optional[str] = Query(None),
    sort_by: Optional[str] = Query(None),
    sort_dir: Optional[str] = Query("desc"),
    db: Session = Depends(get_db),
):
    """Export purchase orders as a styled Excel file respecting all active filters."""
    import io
    import json as _json
    from datetime import datetime
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    orders = crud.get_orders(
        db, skip=0, limit=100_000,
        catalogo=catalogo, categoria=categoria,
        estado_orden=estado_orden, search=search,
        entidad=entidad, proveedor=proveedor,
        sort_by=sort_by, sort_dir=sort_dir,
    )

    # ── Column definitions: (display label, model attr, column width) ────
    COLS = [
        ("Orden Electrónica",        "orden_electronica",    24),
        ("Nro. Orden Física",         "nro_orden_fisica",     20),
        ("Fecha Publicación",         "fecha_publicacion",    18),
        ("Entidad",                   "nombre_entidad",       40),
        ("Proveedor",                 "nombre_proveedor",     40),
        ("Catálogo",                  "catalogo",             22),
        ("Categoría",                 "categoria",            26),
        ("Estado",                    "estado_orden",         14),
        ("Total con IGV (S/)",        "monto_total",          19),
        ("Productos  (P/N · Precio unit. · Subtotal s/IGV)", "nro_parte", 60),
    ]

    # ── Styles ─────────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1E3A8A")
    TITLE_FILL = PatternFill("solid", fgColor="0F172A")
    ALT_FILL   = PatternFill("solid", fgColor="EFF6FF")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    TITLE_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    BOLD_GREEN = Font(bold=True, color="065F46", name="Calibri", size=9)
    BASE_FONT  = Font(name="Calibri", size=9)
    thin       = Side(style="thin", color="CBD5E1")
    BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    WRAP       = Alignment(wrap_text=True, vertical="top")
    TOP        = Alignment(vertical="top")
    R_TOP      = Alignment(horizontal="right", vertical="top")
    MONEY_FMT  = "S/ #,##0.00"

    wb = Workbook()
    ws = wb.active
    ws.title = "Órdenes de Compra"

    # ── Title ────────────────────────────────────────────────────────────
    filters_info = []
    if proveedor:    filters_info.append(f"Proveedor: {proveedor}")
    if entidad:      filters_info.append(f"Entidad: {entidad}")
    if catalogo:     filters_info.append(f"Catálogo: {catalogo}")
    if estado_orden: filters_info.append(f"Estado: {estado_orden}")
    if search:       filters_info.append(f"Búsqueda: {search}")

    title_txt = "Órdenes de Compra — CEAM AUDITOR"
    if filters_info:
        title_txt += f"  |  Filtros: {' · '.join(filters_info)}"
    title_txt += f"  |  {len(orders):,} registros  |  Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    ncols = len(COLS)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    tc = ws["A1"]
    tc.value = title_txt
    tc.fill = TITLE_FILL
    tc.font = TITLE_FONT
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # ── Headers ──────────────────────────────────────────────────────────
    for ci, (label, _, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 28

    # ── Data ─────────────────────────────────────────────────────────────
    for ri, order in enumerate(orders, start=3):
        is_alt = ri % 2 == 0
        for ci, (label, field, _) in enumerate(COLS, start=1):
            raw = getattr(order, field, None)

            if field == "nro_parte":
                try:
                    prods = _json.loads(raw or "[]")
                    if isinstance(prods, list) and prods:
                        lines = []
                        for p in prods:
                            nro = p.get("nro_parte") or "—"
                            pu  = p.get("precio_unitario")
                            pt  = p.get("total")
                            pu_s = f"S/ {float(pu):,.2f}" if pu is not None else "—"
                            pt_s = f"S/ {float(pt):,.2f}" if pt is not None else "—"
                            lines.append(f"• {nro}   unit.: {pu_s}   sub s/IGV: {pt_s}")
                        val = "\n".join(lines)
                    else:
                        val = str(raw) if raw else "—"
                except Exception:
                    val = str(raw) if raw else "—"
            elif field == "monto_total" and raw is not None:
                val = float(raw)
            elif field == "fecha_publicacion" and raw is not None:
                val = str(raw)
            else:
                val = raw

            cell = ws.cell(row=ri, column=ci, value=val)
            if is_alt:
                cell.fill = ALT_FILL
            cell.border = BORDER

            if field == "monto_total":
                cell.font = BOLD_GREEN
                cell.number_format = MONEY_FMT
                cell.alignment = R_TOP
            elif field == "nro_parte":
                cell.font = BASE_FONT
                cell.alignment = WRAP
            else:
                cell.font = BASE_FONT
                cell.alignment = TOP

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    parts = [x[:20].replace(" ", "_") for x in [proveedor, entidad] if x]
    slug  = "_".join(parts) if parts else "todas"
    fname = f"ordenes_{slug}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/{order_id}", response_model=PurchaseOrderResponse)
def get_order(order_id: int, db: Session = Depends(get_db)):
    order = crud.get_order(db, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


@router.post("/", response_model=PurchaseOrderResponse, status_code=201)
def create_order(payload: PurchaseOrderCreate, db: Session = Depends(get_db)):
    existing = crud.get_order_by_electronica(db, payload.orden_electronica)
    if existing:
        raise HTTPException(status_code=409, detail="Order number already exists")
    return crud.create_order(db, payload)


@router.delete("/{order_id}", status_code=204)
def delete_order(order_id: int, db: Session = Depends(get_db)):
    success = crud.delete_order(db, order_id)
    if not success:
        raise HTTPException(status_code=404, detail="Order not found")
